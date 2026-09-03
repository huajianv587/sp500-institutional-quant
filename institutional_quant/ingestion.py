from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path

import pandas as pd

from .config import Settings
from .schemas import DataQualityIssue, DatasetKind, ImportResult, Severity
from .storage import Store

ALIASES: dict[str, tuple[str, ...]] = {
    "company_id": (
        "company_id",
        "companyid",
        "iq_company_id",
        "spgiq_company_id",
        "entity_id",
        "sp_entity_id",
        "spciq_id",
    ),
    "ticker": (
        "ticker",
        "trading_symbol",
        "symbol",
        "exchange_ticker",
        "sp_exchange_ticker",
    ),
    "company_name": (
        "company_name",
        "company",
        "name",
        "entity_name",
        "sp_entity_name",
    ),
    "sector": ("sector", "gics_sector", "industry_sector", "iq_sector"),
    "currency": ("currency", "reporting_currency"),
    "effective_at": (
        "effective_at",
        "speffectivedate",
        "sp_effective_date",
        "filing_date",
        "financial_filing_date",
        "iq_finl_filing_date",
        "available_at",
        "data_available_date",
    ),
    "as_of_date": ("as_of_date", "asofdate", "snapshot_date", "export_date"),
    "member_from": ("member_from", "membership_start", "start_date", "effective_from"),
    "member_to": ("member_to", "membership_end", "end_date", "effective_to"),
    "index_code": ("index_code", "index", "index_name"),
    "period_end": ("period_end", "periodend", "fiscal_period_end", "iq_period_end"),
    "period_type": ("period_type", "periodtype", "frequency"),
    "metric": ("metric", "data_item", "item", "field"),
    "value": ("value", "metric_value", "data_value"),
    "unit": ("unit", "units"),
    "fiscal_period": (
        "fiscal_period",
        "fiscal_period_end",
        "estimate_period",
        "sp_eps_norm_date_est",
    ),
    "valid_to": ("valid_to", "sptodate", "sp_to_date"),
    "price_date": ("price_date", "date", "trading_date"),
    "open": ("open", "open_price"),
    "high": ("high", "high_price"),
    "low": ("low", "low_price"),
    "close": ("close", "close_price"),
    "adjusted_close": ("adjusted_close", "adj_close", "adjustedclose"),
    "return_1d": ("return_1d", "return_1d_pct", "price_change_1d"),
    "return_1w": ("return_1w", "return_1w_pct", "price_change_1w"),
    "return_1m": ("return_1m", "return_1m_pct", "price_change_1m"),
    "volume": ("volume", "trading_volume"),
    "source": ("source", "price_source"),
    "institutional_pct": ("institutional_pct", "institutional_ownership_pct"),
    "institutional_change": ("institutional_change", "institutional_ownership_change"),
    "transaction_date": ("transaction_date", "trade_date"),
    "transaction_type": ("transaction_type", "type", "trade_type"),
    "shares": ("shares", "share_count"),
}

# Capital IQ exports use keyfield labels while hand-curated fixtures and other
# enterprise feeds often use plain-English names.  The raw header is retained in
# the immutable source archive; this map gives the analytical layer one stable
# metric vocabulary.
CANONICAL_METRIC_ALIASES: dict[str, str] = {
    "iq_total_rev": "revenue",
    "iq_total_revenue": "revenue",
    "total_revenue": "revenue",
    "iq_net_income": "net_income",
    "iq_net_inc_parent": "net_income",
    "iq_fcf": "free_cash_flow",
    "iq_free_cash_flow": "free_cash_flow",
    "iq_cash_from_ops": "operating_cash_flow",
    "iq_cash_from_oper": "operating_cash_flow",
    "iq_cash_oper": "operating_cash_flow",
    "iq_operating_cash_flow": "operating_cash_flow",
    "cash_from_operations": "operating_cash_flow",
    "iq_ebitda": "ebitda",
    "iq_nopat": "nopat",
    "iq_invested_capital": "invested_capital",
    "iq_gross_profit": "gross_profit",
    "iq_gp": "gross_profit",
    "iq_total_assets": "total_assets",
    "iq_total_debt": "total_debt",
    "iq_total_equity": "total_equity",
    "iq_capex": "capital_expenditure",
    "iq_net_debt": "net_debt",
    "iq_market_cap": "market_cap",
    "sp_marketcap": "market_cap",
    "market_capitalization": "market_cap",
    "iq_tev": "enterprise_value",
    "iq_total_enterprise_value": "enterprise_value",
    "total_enterprise_value": "enterprise_value",
    "iq_diluted_eps": "eps",
    "iq_eps_diluted": "eps",
    "diluted_eps": "eps",
    "iq_operating_margin": "operating_margin",
    "iq_oper_margin": "operating_margin",
    "iq_total_rev_1yr_ann_growth": "revenue_growth",
    "iq_gross_margin": "gross_margin",
    "iq_ebitda_margin": "ebitda_margin",
    "iq_ni_margin": "net_income_margin",
    "iq_tev_ebitda": "tev_ebitda",
    "iq_pe": "price_to_earnings",
    "iq_pbv_x": "price_to_book",
    "iq_roa": "return_on_assets",
    "iq_roc": "roic",
    "iq_roe": "return_on_equity",
    "sp_norm_eps_act_or_est": "eps_estimate",
    "sp_normalized_eps_actual_estimate": "eps_estimate",
    "sp_eps_norm_est": "eps_estimate",
    "sp_revenue_estimate": "revenue_estimate",
    "sp_rev_est": "revenue_estimate",
    "sp_eps_norm_est_num_analysts_month": "eps_analyst_count_1m",
    "sp_eps_norm_est_up_month": "eps_up_revisions_1m",
    "sp_eps_norm_est_down_month": "eps_down_revisions_1m",
    "sp_eps_norm_est_up_3month": "eps_up_revisions_3m",
    "sp_eps_norm_est_down_3month": "eps_down_revisions_3m",
    "iq_eps_revision_1m": "eps_revision_1m",
    "eps_revision_1_month": "eps_revision_1m",
    "iq_eps_revision_3m": "eps_revision_3m",
    "eps_revision_3_month": "eps_revision_3m",
    "iq_estimate_surprise": "estimate_surprise",
}

CIQ_PERCENTAGE_POINT_METRICS = {
    "iq_total_rev_1yr_ann_growth",
    "iq_gross_margin",
    "iq_ebitda_margin",
    "iq_ni_margin",
    "iq_roa",
    "iq_roc",
    "iq_roe",
}

# Capital IQ's LTM valuation multiples are themselves observable on the
# requested as-of date.  They do not share a fiscal filing timestamp with the
# FY statement fields, so pairing them to an FY0 filing date would be
# semantically wrong.  Preserve their LTM period type while using the vendor
# snapshot date as both the analytical period anchor and availability date.
CIQ_AS_OF_FUNDAMENTAL_METRICS = {
    "iq_pe",
    "iq_tev_ebitda",
}


def _column_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


@dataclass(frozen=True)
class Contract:
    table: str
    required: tuple[str, ...]
    optional: tuple[str, ...] = ()
    date_columns: tuple[str, ...] = ()
    datetime_columns: tuple[str, ...] = ("effective_at",)
    defaults: dict[str, object] | None = None


CONTRACTS: dict[DatasetKind, Contract] = {
    DatasetKind.INSTRUMENTS: Contract(
        "instruments",
        ("company_id", "ticker", "company_name", "sector", "effective_at", "as_of_date"),
        ("currency",),
        date_columns=("as_of_date",),
        defaults={"currency": "USD"},
    ),
    DatasetKind.INDEX_MEMBERSHIP: Contract(
        "index_membership",
        (
            "company_id",
            "ticker",
            "member_from",
            "effective_at",
            "as_of_date",
        ),
        ("member_to", "index_code"),
        date_columns=("member_from", "member_to", "as_of_date"),
        defaults={"index_code": "SP500"},
    ),
    DatasetKind.FUNDAMENTALS: Contract(
        "fundamentals",
        (
            "company_id",
            "ticker",
            "period_end",
            "period_type",
            "effective_at",
            "as_of_date",
            "metric",
            "value",
        ),
        ("unit",),
        date_columns=("period_end", "as_of_date"),
    ),
    DatasetKind.ESTIMATES: Contract(
        "estimates",
        (
            "company_id",
            "ticker",
            "fiscal_period",
            "effective_at",
            "as_of_date",
            "metric",
            "value",
        ),
        ("valid_to", "unit"),
        date_columns=("fiscal_period", "as_of_date"),
        datetime_columns=("effective_at", "valid_to"),
    ),
    DatasetKind.PRICES: Contract(
        "prices",
        (
            "company_id",
            "ticker",
            "price_date",
            "open",
            "high",
            "low",
            "close",
            "adjusted_close",
            "effective_at",
            "as_of_date",
        ),
        ("volume", "source"),
        date_columns=("price_date", "as_of_date"),
        defaults={"source": "capital_iq"},
    ),
    DatasetKind.MARKET_RETURNS: Contract(
        "market_returns",
        ("company_id", "ticker", "as_of_date", "effective_at", "return_1d", "return_1w", "return_1m"),
        (),
        date_columns=("as_of_date",),
        datetime_columns=("effective_at",),
    ),
    DatasetKind.OWNERSHIP: Contract(
        "ownership",
        ("company_id", "ticker", "effective_at", "as_of_date"),
        ("institutional_pct", "institutional_change"),
        date_columns=("as_of_date",),
    ),
    DatasetKind.INSIDER_TRANSACTIONS: Contract(
        "insider_transactions",
        (
            "company_id",
            "ticker",
            "transaction_date",
            "effective_at",
            "as_of_date",
            "transaction_type",
            "shares",
        ),
        ("value",),
        date_columns=("transaction_date", "as_of_date"),
    ),
}


TABLE_COLUMNS: dict[str, list[str]] = {
    "instruments": [
        "company_id",
        "ticker",
        "company_name",
        "sector",
        "currency",
        "effective_at",
        "as_of_date",
        "source_file_id",
        "ingested_at",
    ],
    "index_membership": [
        "company_id",
        "ticker",
        "index_code",
        "member_from",
        "member_to",
        "effective_at",
        "as_of_date",
        "source_file_id",
        "ingested_at",
    ],
    "fundamentals": [
        "company_id",
        "ticker",
        "period_end",
        "period_type",
        "effective_at",
        "as_of_date",
        "metric",
        "value",
        "unit",
        "source_file_id",
        "ingested_at",
    ],
    "estimates": [
        "company_id",
        "ticker",
        "fiscal_period",
        "effective_at",
        "valid_to",
        "as_of_date",
        "metric",
        "value",
        "unit",
        "source_file_id",
        "ingested_at",
    ],
    "prices": [
        "company_id",
        "ticker",
        "price_date",
        "open",
        "high",
        "low",
        "close",
        "adjusted_close",
        "volume",
        "source",
        "effective_at",
        "as_of_date",
        "source_file_id",
        "ingested_at",
    ],
    "market_returns": [
        "company_id",
        "ticker",
        "as_of_date",
        "effective_at",
        "return_1d",
        "return_1w",
        "return_1m",
        "source_file_id",
        "ingested_at",
    ],
    "ownership": [
        "company_id",
        "ticker",
        "effective_at",
        "as_of_date",
        "institutional_pct",
        "institutional_change",
        "source_file_id",
        "ingested_at",
    ],
    "insider_transactions": [
        "company_id",
        "ticker",
        "transaction_date",
        "effective_at",
        "as_of_date",
        "transaction_type",
        "shares",
        "value",
        "source_file_id",
        "ingested_at",
    ],
}


class CapitalIQImporter:
    def __init__(self, store: Store, settings: Settings):
        self.store = store
        self.settings = settings

    @staticmethod
    def sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def read(path: Path) -> pd.DataFrame:
        frame, _ = CapitalIQImporter.read_with_metadata(path)
        return frame

    @staticmethod
    def read_with_metadata(path: Path) -> tuple[pd.DataFrame, dict[str, object]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return pd.read_csv(path), {}
        if suffix in {".xlsx", ".xlsm", ".xls"}:
            preview = pd.read_excel(path, header=None, nrows=25)
            header_row = CapitalIQImporter._detect_header_row(preview)
            frame = pd.read_excel(path, header=header_row)
            metadata = CapitalIQImporter._excel_formula_metadata(path, frame.columns)
            return frame.dropna(how="all").reset_index(drop=True), metadata
        raise ValueError("Capital IQ import supports CSV, XLSX, XLSM, and XLS")

    @staticmethod
    def _parse_spg_formula_metadata(formulas: Iterable[str]) -> dict[str, object]:
        as_of_dates: set[str] = set()
        period_codes: set[str] = set()
        period_codes_by_keyfield: dict[str, set[str]] = {}
        for formula in formulas:
            if "SPGLabel" not in formula:
                continue
            for month, day, year in re.findall(
                r'"(?:<>|<=|>=)?(\d{1,2})/(\d{1,2})/(\d{4})"', formula
            ):
                as_of_dates.add(date(int(year), int(month), int(day)).isoformat())
            match = re.search(
                r'SPGLabel\([^,]+,\s*([^,]+),\s*"([A-Za-z]+(?:[+-]?\d+)?)"',
                formula,
            )
            if match:
                keyfield = match.group(1).strip()
                period_code = match.group(2).upper()
                period_codes.add(period_code)
                period_codes_by_keyfield.setdefault(keyfield, set()).add(period_code)
        metadata: dict[str, object] = {}
        if len(as_of_dates) == 1:
            metadata["embedded_as_of_date"] = next(iter(as_of_dates))
        elif as_of_dates:
            metadata["embedded_as_of_dates"] = sorted(as_of_dates)
        if period_codes:
            metadata["embedded_period_codes"] = sorted(period_codes)
            metadata["embedded_period_codes_by_keyfield"] = {
                keyfield: sorted(codes)
                for keyfield, codes in sorted(period_codes_by_keyfield.items())
            }
        return metadata

    @staticmethod
    def _excel_formula_metadata(
        path: Path, actual_columns: Iterable[object] | None = None
    ) -> dict[str, object]:
        if path.suffix.lower() == ".xls":
            return {}
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            formulas = [
                value
                for row in sheet.iter_rows(min_row=1, max_row=12, values_only=True)
                for value in row
                if isinstance(value, str) and value.startswith("=")
            ]
            metadata = CapitalIQImporter._parse_spg_formula_metadata(formulas)
            preview = pd.read_excel(path, header=None, nrows=25)
            header_row = CapitalIQImporter._detect_header_row(preview)
            table_row = next(
                (
                    row_number
                    for row_number in range(1, min(sheet.max_row, 12) + 1)
                    if any(
                        isinstance(sheet.cell(row_number, column).value, str)
                        and "SPGTable" in sheet.cell(row_number, column).value
                        for column in range(1, sheet.max_column + 1)
                    )
                ),
                None,
            )
            label_row = table_row + 1 if table_row is not None else None
            if label_row is None:
                candidate_rows = range(1, min(header_row + 1, sheet.max_row) + 1)
                label_counts = {
                    row_number: sum(
                        isinstance(sheet.cell(row_number, column).value, str)
                        and "SPGLabel" in sheet.cell(row_number, column).value
                        for column in range(1, sheet.max_column + 1)
                    )
                    for row_number in candidate_rows
                }
                if label_counts and max(label_counts.values()) > 0:
                    label_row = max(label_counts, key=label_counts.get)
            if label_row is not None and label_row <= sheet.max_row:
                by_header: dict[str, set[str]] = {}
                embedded_columns: list[dict[str, str]] = []
                parsed_columns = list(actual_columns) if actual_columns is not None else []
                for column in range(1, min(sheet.max_column, preview.shape[1]) + 1):
                    header = preview.iloc[header_row, column - 1]
                    formula = sheet.cell(label_row, column).value
                    if pd.isna(header) or not isinstance(formula, str) or "SPGLabel" not in formula:
                        continue
                    match = re.search(r"SPGLabel\(\s*[^,]+,\s*([^,]+)", formula)
                    quoted = re.findall(r'"([^"]*)"', formula)
                    period_code = quoted[0].upper() if quoted else ""
                    as_of_token = next(
                        (
                            token
                            for token in quoted[1:]
                            if token.lower() == "current"
                            or re.fullmatch(r"(?:<>|<=|>=)?\d{1,2}/\d{1,2}/\d{4}", token)
                        ),
                        None,
                    )
                    entry = {
                        "column": (
                            str(parsed_columns[column - 1])
                            if column - 1 < len(parsed_columns)
                            else str(header)
                        ),
                        "source_header": str(header),
                        "header_key": _column_key(header),
                    }
                    if match:
                        entry["keyfield"] = match.group(1).strip()
                    if period_code:
                        entry["period_code"] = period_code
                        by_header.setdefault(_column_key(header), set()).add(period_code)
                    if as_of_token:
                        date_match = re.fullmatch(
                            r"(?:<>|<=|>=)?(\d{1,2})/(\d{1,2})/(\d{4})", as_of_token
                        )
                        if date_match:
                            month, day, year = (int(value) for value in date_match.groups())
                            entry["as_of_date"] = date(year, month, day).isoformat()
                        else:
                            entry["as_of"] = "current"
                    embedded_columns.append(entry)
                if by_header:
                    metadata["embedded_period_codes_by_header"] = {
                        header: sorted(codes) for header, codes in sorted(by_header.items())
                    }
                if embedded_columns:
                    metadata["embedded_columns"] = embedded_columns
            return metadata
        finally:
            workbook.close()

    @staticmethod
    def _detect_header_row(preview: pd.DataFrame) -> int:
        """Locate the real table header below Capital IQ's Excel formula preamble."""
        known = {alias for aliases in ALIASES.values() for alias in aliases} | set(ALIASES)
        identity = {
            "company_id",
            "companyid",
            "iq_company_id",
            "spgiq_company_id",
            "entity_id",
            "sp_entity_id",
            "spciq_id",
            "ticker",
            "exchange_ticker",
            "sp_exchange_ticker",
            "metric",
            "price_date",
        }
        best_row = 0
        best_score = 0
        for row_number, row in preview.iterrows():
            keys = {
                _column_key(value)
                for value in row.tolist()
                if pd.notna(value) and str(value).strip()
            }
            score = len(keys & known)
            if score > best_score and keys & identity:
                best_row = int(row_number)
                best_score = score
        return best_row if best_score >= 2 else 0

    @staticmethod
    def _rename_columns(frame: pd.DataFrame) -> pd.DataFrame:
        keyed = {_column_key(column): column for column in frame.columns}
        renames: dict[str, str] = {}
        for canonical, aliases in ALIASES.items():
            for alias in aliases:
                if alias in keyed:
                    renames[keyed[alias]] = canonical
                    break
        return frame.rename(columns=renames)

    @staticmethod
    def _rename_market_return_columns(frame: pd.DataFrame) -> pd.DataFrame:
        """Map CIQ's repeated ``SP_PRICE_CHANGE`` columns by exported order.

        pandas disambiguates duplicate workbook headers with ``.1``/``.2``;
        positional mapping is therefore the only stable interpretation of the
        1D, 1W and 1M columns in a Results As Values export.
        """
        columns = list(frame.columns)
        candidates = [
            column
            for column in columns
            if _column_key(column).startswith("price_change")
            or _column_key(column).startswith("sp_price_change")
        ]
        if len(candidates) < 3:
            # Some CSV exports retain the keyfield labels as headers.
            candidates = [column for column in columns if "price_change" in _column_key(column)]
        if len(candidates) < 3:
            return frame
        renames = {
            candidates[0]: "return_1d",
            candidates[1]: "return_1w",
            candidates[2]: "return_1m",
        }
        return frame.rename(columns=renames)

    @staticmethod
    def _normalize_identifiers(frame: pd.DataFrame) -> pd.DataFrame:
        normalized = frame.copy()
        if "company_id" in normalized.columns:
            normalized["company_id"] = normalized["company_id"].map(
                lambda value: (
                    None
                    if pd.isna(value)
                    else str(int(value))
                    if isinstance(value, float) and value.is_integer()
                    else str(value).strip()
                )
            )
        if "ticker" in normalized.columns:
            normalized["ticker"] = (
                normalized["ticker"].astype("string").str.strip().str.rsplit(":", n=1).str[-1]
            )
        if "company_name" in normalized.columns:
            normalized["company_name"] = (
                normalized["company_name"]
                .astype("string")
                .str.strip()
                .str.replace(r"\s+\([A-Za-z0-9._-]+:[^)]+\)$", "", regex=True)
            )
        if "sector" in normalized.columns:
            normalized["sector"] = normalized["sector"].astype("string").str.strip()
        return normalized

    @staticmethod
    def _drop_parameter_rows(frame: pd.DataFrame) -> tuple[pd.DataFrame, int]:
        if "company_id" not in frame.columns:
            return frame, 0
        identity_columns = [
            column for column in ("company_id", "ticker", "company_name") if column in frame.columns
        ]
        missing_identity = frame[identity_columns].isna().all(axis=1)

        def is_parameter_row(row: pd.Series) -> bool:
            values = [
                str(value).strip()
                for value in row.tolist()
                if pd.notna(value) and str(value).strip()
            ]
            return bool(values) and all(
                re.fullmatch(r"[A-Za-z]+(?:[+-]?\d+)?", value) for value in values
            )

        parameter_rows = missing_identity & frame.apply(is_parameter_row, axis=1)
        return frame.loc[~parameter_rows].copy(), int(parameter_rows.sum())

    @staticmethod
    def _derive_estimate_fiscal_period(
        frame: pd.DataFrame, source_metadata: dict[str, object]
    ) -> tuple[pd.DataFrame, str | None]:
        """Derive a dated estimate target from CIQ's relative FY/FQ period code.

        Capital IQ estimate fields identify the target as FY+1/FQ+1 in the
        SPGLabel formula.  A companion point-in-time Period Ended field supplies
        each company's FY0/FQ0 date.  The relative shift is deterministic and
        keeps the exported file self-contained.
        """
        if "fiscal_period" in frame.columns or "period_end" not in frame.columns:
            return frame, None
        raw_codes = source_metadata.get("embedded_period_codes")
        if not isinstance(raw_codes, list):
            return frame, None
        parsed: list[tuple[str, int, str]] = []
        for raw_code in raw_codes:
            match = re.fullmatch(r"([A-Za-z]+)([+-]?\d+)", str(raw_code))
            if not match:
                continue
            prefix = match.group(1).upper()
            offset = int(match.group(2))
            if offset != 0 and prefix in {"FY", "FQ"}:
                parsed.append((prefix, offset, str(raw_code).upper()))
        if len(parsed) != 1:
            return frame, None

        prefix, offset, period_code = parsed[0]
        months = offset * (12 if prefix == "FY" else 3)
        derived = frame.copy()
        period_end = pd.to_datetime(derived["period_end"], errors="coerce")
        derived["fiscal_period"] = period_end + pd.DateOffset(months=months)
        return derived, period_code

    @staticmethod
    def _expand_historical_estimate_snapshots(
        frame: pd.DataFrame, source_metadata: dict[str, object]
    ) -> pd.DataFrame:
        """Convert repeated CIQ as-of columns into dated estimate observations."""
        raw_columns = source_metadata.get("embedded_columns")
        if not isinstance(raw_columns, list):
            return frame
        dated = [
            item
            for item in raw_columns
            if isinstance(item, dict)
            and isinstance(item.get("column"), str)
            and isinstance(item.get("as_of_date"), str)
            and item["column"] in frame.columns
        ]
        dated_header_keys = {str(item.get("header_key")) for item in dated}
        has_current_sibling = any(
            isinstance(item, dict)
            and item.get("as_of") == "current"
            and str(item.get("header_key")) in dated_header_keys
            for item in raw_columns
        )
        if len({str(item["as_of_date"]) for item in dated}) < 2 and not has_current_sibling:
            return frame

        fiscal_period_aliases = set(ALIASES["fiscal_period"]) | set(ALIASES["period_end"])
        period_columns = [
            item for item in dated if str(item.get("header_key")) in fiscal_period_aliases
        ]
        identity_canonicals = {
            "company_id",
            "ticker",
            "company_name",
            "sector",
            "currency",
            "unit",
            "valid_to",
        }
        identity_aliases = {
            alias for canonical in identity_canonicals for alias in ALIASES.get(canonical, ())
        }
        identity_columns = [
            str(column) for column in frame.columns if _column_key(column) in identity_aliases
        ]
        observation_rows = (
            frame[identity_columns].notna().any(axis=1)
            if identity_columns
            else pd.Series(True, index=frame.index)
        )
        source_rows = frame.loc[observation_rows]
        value_columns = [
            item
            for item in dated
            if item not in period_columns and str(item.get("header_key")) not in identity_aliases
        ]
        if not value_columns:
            return frame

        observations: list[pd.DataFrame] = []
        for item in value_columns:
            as_of_date = str(item["as_of_date"])
            period_code = str(item.get("period_code", ""))
            companion = next(
                (
                    candidate
                    for candidate in period_columns
                    if candidate.get("as_of_date") == as_of_date
                    and (
                        not period_code
                        or not candidate.get("period_code")
                        or candidate.get("period_code") == period_code
                    )
                ),
                None,
            )
            selected = source_rows[identity_columns].copy()
            selected["as_of_date"] = as_of_date
            selected["effective_at"] = as_of_date
            selected["fiscal_period"] = (
                source_rows[str(companion["column"])] if companion is not None else None
            )
            selected["metric"] = str(item.get("source_header", item["column"]))
            selected["value"] = source_rows[str(item["column"])]
            observations.append(selected)

        source_metadata["historical_snapshot_dates"] = sorted(
            {str(item["as_of_date"]) for item in value_columns}
        )
        source_metadata["estimate_availability_source"] = "embedded_spg_as_of_date_by_column"
        source_metadata["estimate_fiscal_period_policy"] = (
            "pair_same_as_of_and_period_code_companion"
        )
        return pd.concat(observations, ignore_index=True)

    @staticmethod
    def _expand_historical_fundamental_snapshots(
        frame: pd.DataFrame, source_metadata: dict[str, object]
    ) -> pd.DataFrame:
        """Pair each historical fundamental value with its period and filing date."""
        raw_columns = source_metadata.get("embedded_columns")
        if not isinstance(raw_columns, list):
            return frame
        dated = [
            item
            for item in raw_columns
            if isinstance(item, dict)
            and isinstance(item.get("column"), str)
            and isinstance(item.get("as_of_date"), str)
            and item["column"] in frame.columns
        ]
        dated_header_keys = {str(item.get("header_key")) for item in dated}
        has_current_sibling = any(
            isinstance(item, dict)
            and item.get("as_of") == "current"
            and str(item.get("header_key")) in dated_header_keys
            for item in raw_columns
        )
        if len({str(item["as_of_date"]) for item in dated}) < 2 and not has_current_sibling:
            return frame

        period_aliases = set(ALIASES["period_end"])
        filing_aliases = set(ALIASES["effective_at"])
        period_columns = [item for item in dated if str(item.get("header_key")) in period_aliases]
        filing_columns = [item for item in dated if str(item.get("header_key")) in filing_aliases]
        identity_canonicals = {
            "company_id",
            "ticker",
            "company_name",
            "sector",
            "currency",
            "unit",
        }
        identity_aliases = {
            alias for canonical in identity_canonicals for alias in ALIASES.get(canonical, ())
        }
        identity_columns = [
            str(column) for column in frame.columns if _column_key(column) in identity_aliases
        ]
        observation_rows = (
            frame[identity_columns].notna().any(axis=1)
            if identity_columns
            else pd.Series(True, index=frame.index)
        )
        source_rows = frame.loc[observation_rows]
        metadata_aliases = identity_aliases | period_aliases | filing_aliases
        value_columns = [
            item for item in dated if str(item.get("header_key")) not in metadata_aliases
        ]
        if not value_columns:
            return frame

        def companion_for(
            candidates: list[dict[str, object]], *, as_of_date: str, period_code: str
        ) -> dict[str, object] | None:
            return next(
                (
                    candidate
                    for candidate in candidates
                    if candidate.get("as_of_date") == as_of_date
                    and (
                        not period_code
                        or not candidate.get("period_code")
                        or candidate.get("period_code") == period_code
                    )
                ),
                None,
            )

        observations: list[pd.DataFrame] = []
        as_of_observed_metrics: set[str] = set()
        for item in value_columns:
            as_of_date = str(item["as_of_date"])
            period_code = str(item.get("period_code", ""))
            period = companion_for(period_columns, as_of_date=as_of_date, period_code=period_code)
            filing = companion_for(filing_columns, as_of_date=as_of_date, period_code=period_code)
            metric_key = _column_key(str(item.get("source_header", item["column"])))
            is_as_of_observed = metric_key in CIQ_AS_OF_FUNDAMENTAL_METRICS
            if is_as_of_observed:
                as_of_observed_metrics.add(metric_key)
            selected = source_rows[identity_columns].copy()
            selected["as_of_date"] = as_of_date
            selected["period_end"] = (
                as_of_date
                if is_as_of_observed
                else source_rows[str(period["column"])]
                if period is not None
                else None
            )
            selected["effective_at"] = (
                as_of_date
                if is_as_of_observed
                else source_rows[str(filing["column"])]
                if filing is not None
                else None
            )
            selected["metric"] = str(item.get("source_header", item["column"]))
            selected["value"] = source_rows[str(item["column"])]
            observations.append(selected)

        source_metadata["historical_snapshot_dates"] = sorted(
            {str(item["as_of_date"]) for item in value_columns}
        )
        source_metadata["fundamental_availability_source"] = (
            "same_as_of_spg_financial_filing_date_by_column"
        )
        source_metadata["fundamental_period_policy"] = "pair_same_as_of_and_period_code_companion"
        if as_of_observed_metrics:
            source_metadata["as_of_observed_fundamental_metrics"] = sorted(as_of_observed_metrics)
            source_metadata["as_of_observed_metric_policy"] = (
                "period_end_and_effective_at_equal_embedded_spg_as_of_date"
            )
        return pd.concat(observations, ignore_index=True)

    @staticmethod
    def _wide_to_long(frame: pd.DataFrame, dataset: DatasetKind) -> pd.DataFrame:
        if dataset not in {DatasetKind.FUNDAMENTALS, DatasetKind.ESTIMATES}:
            return frame
        if {"metric", "value"}.issubset(frame.columns):
            return frame
        identity = {
            "company_id",
            "ticker",
            "company_name",
            "period_end",
            "period_type",
            "fiscal_period",
            "effective_at",
            "valid_to",
            "as_of_date",
            "unit",
            "sector",
        }
        metadata_columns = {"mi_primary_industry", "iq_primary_industry"}
        value_columns = [
            column
            for column in frame.columns
            if column not in identity and _column_key(column) not in metadata_columns
        ]
        if not value_columns:
            return frame
        return frame.melt(
            id_vars=[column for column in frame.columns if column in identity],
            value_vars=value_columns,
            var_name="metric",
            value_name="value",
        )

    @staticmethod
    def _derive_fundamental_period_types(
        frame: pd.DataFrame,
        source_metadata: dict[str, object],
        *,
        current_snapshot_as_of: date | None = None,
        current_snapshot_effective_at: datetime | None = None,
    ) -> pd.DataFrame:
        """Assign each wide CIQ metric its own FY/LTM/current period semantics."""
        if "metric" not in frame.columns:
            return frame
        derived = frame.copy()
        header_periods = source_metadata.get("embedded_period_codes_by_header")
        period_by_header = header_periods if isinstance(header_periods, dict) else {}

        def period_type(metric: object) -> str | None:
            codes = period_by_header.get(_column_key(metric), [])
            if not isinstance(codes, list) or len(codes) != 1:
                return None
            match = re.match(r"[A-Za-z]+", str(codes[0]))
            return match.group(0).upper() if match else None

        inferred = derived["metric"].map(period_type)
        if "period_type" not in derived.columns:
            derived["period_type"] = inferred
        else:
            derived["period_type"] = derived["period_type"].fillna(inferred)

        if current_snapshot_as_of is not None:
            current_metric_keys = {"sp_marketcap", "iq_tev"}
            current_mask = derived["metric"].map(_column_key).isin(current_metric_keys)
            if current_mask.any():
                derived.loc[current_mask, "period_type"] = "CURRENT"
                derived["period_end"] = derived["period_end"].astype(object)
                derived.loc[current_mask, "period_end"] = current_snapshot_as_of
                if current_snapshot_effective_at is not None:
                    derived["effective_at"] = derived["effective_at"].astype(object)
                    derived.loc[current_mask, "effective_at"] = current_snapshot_effective_at
        return derived

    def import_file(
        self,
        path: str | Path,
        dataset: DatasetKind,
        *,
        current_snapshot_as_of: date | None = None,
        current_snapshot_effective_at: datetime | None = None,
    ) -> ImportResult:
        current_snapshot = any(
            value is not None for value in (current_snapshot_as_of, current_snapshot_effective_at)
        )
        if current_snapshot:
            allowed_current = {
                DatasetKind.INSTRUMENTS,
                DatasetKind.FUNDAMENTALS,
                DatasetKind.ESTIMATES,
                DatasetKind.MARKET_RETURNS,
            }
            if dataset not in allowed_current:
                raise ValueError(
                    "Current-snapshot timestamps are permitted only for instruments, "
                    "fundamentals, and estimates"
                )
            if current_snapshot_as_of is None or current_snapshot_effective_at is None:
                raise ValueError(
                    "Both current_snapshot_as_of and current_snapshot_effective_at are required"
                )
        if self.store.is_cloud and not self.settings.ciq_cloud_storage_confirmed:
            raise RuntimeError(
                "Set CIQ_CLOUD_STORAGE_CONFIRMED=true only after confirming that "
                "your NTU/S&P agreement permits storing Capital IQ values in Supabase"
            )
        source = Path(path)
        if not source.is_file():
            raise FileNotFoundError(source)
        digest = self.sha256(source)
        existing = self.store.source_by_hash(digest)
        if existing:
            return ImportResult(
                source_file_id=str(existing["source_file_id"]),
                dataset=dataset,
                sha256=digest,
                archived_path=str(existing["archived_path"]),
                imported_rows=int(existing["row_count"]),
                idempotent=True,
            )

        source_file_id = f"src_{digest[:24]}"
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", source.name)
        archived = self.settings.raw_data_dir / dataset.value / f"{digest[:12]}_{safe_name}"
        archived.parent.mkdir(parents=True, exist_ok=True)
        # External exFAT/APFS volumes can reject macOS file flags during copy2's
        # copystat phase.  The immutable archive is content-addressed and audited
        # by SHA-256, so copying bytes is the portable and sufficient operation.
        shutil.copyfile(source, archived)

        frame, source_metadata = self.read_with_metadata(archived)
        source_column_keys = {_column_key(column) for column in frame.columns}
        date_only_availability = bool(
            source_column_keys & {"filing_date", "financial_filing_date", "iq_finl_filing_date"}
        )
        if date_only_availability:
            source_metadata["effective_at_granularity"] = "date"
            source_metadata["date_only_availability_policy"] = "end_of_day"
        if dataset is DatasetKind.FUNDAMENTALS:
            frame = self._expand_historical_fundamental_snapshots(frame, source_metadata)
            if source_metadata.get("historical_snapshot_dates"):
                date_only_availability = True
                source_metadata["effective_at_granularity"] = "date"
                source_metadata["date_only_availability_policy"] = "end_of_day"
        elif dataset is DatasetKind.ESTIMATES:
            frame = self._expand_historical_estimate_snapshots(frame, source_metadata)
            if source_metadata.get("historical_snapshot_dates"):
                date_only_availability = True
                source_metadata["effective_at_granularity"] = "date"
                source_metadata["date_only_availability_policy"] = "end_of_day"
        if dataset is DatasetKind.MARKET_RETURNS:
            frame = self._rename_market_return_columns(frame)
        frame = self._rename_columns(frame)
        frame = self._normalize_identifiers(frame)
        if dataset is DatasetKind.MARKET_RETURNS and {"company_id", "ticker"}.issubset(frame.columns):
            # Results As Values places Current/1D/1W/1M parameter rows directly
            # below the keyfield row; they are not observations and should not
            # become fatal timestamp errors.
            frame = frame.loc[frame["company_id"].notna() & frame["ticker"].notna()].copy()
        frame, parameter_rows_dropped = self._drop_parameter_rows(frame)
        if (
            current_snapshot
            and dataset is DatasetKind.FUNDAMENTALS
            and "effective_at" not in frame.columns
        ):
            raise ValueError(
                "A current fundamentals snapshot still requires a source-provided "
                "Financial Filing Date/effective_at column"
            )
        if current_snapshot:
            if "as_of_date" not in frame.columns:
                frame["as_of_date"] = current_snapshot_as_of
            if "effective_at" not in frame.columns:
                frame["effective_at"] = current_snapshot_effective_at
        elif "as_of_date" not in frame.columns and source_metadata.get("embedded_as_of_date"):
            frame["as_of_date"] = source_metadata["embedded_as_of_date"]
        if (
            dataset is DatasetKind.ESTIMATES
            and "effective_at" not in frame.columns
            and source_metadata.get("embedded_as_of_date")
        ):
            frame["effective_at"] = source_metadata["embedded_as_of_date"]
            date_only_availability = True
            source_metadata["effective_at_granularity"] = "date"
            source_metadata["date_only_availability_policy"] = "end_of_day"
            source_metadata["estimate_availability_source"] = "embedded_spg_as_of_date"
        if dataset is DatasetKind.FUNDAMENTALS and "period_type" not in frame.columns:
            period_codes = source_metadata.get("embedded_period_codes")
            prefixes = (
                {re.match(r"[A-Za-z]+", code).group(0) for code in period_codes}
                if isinstance(period_codes, list) and period_codes
                else set()
            )
            if len(prefixes) == 1:
                frame["period_type"] = next(iter(prefixes))
        if dataset is DatasetKind.ESTIMATES:
            frame, estimate_period_code = self._derive_estimate_fiscal_period(
                frame, source_metadata
            )
            if estimate_period_code:
                source_metadata["estimate_period_code"] = estimate_period_code
                source_metadata["estimate_fiscal_period_policy"] = (
                    "shift_companion_period_end_by_spg_period_code"
                )
        frame = self._wide_to_long(frame, dataset)
        if dataset is DatasetKind.FUNDAMENTALS:
            frame = self._derive_fundamental_period_types(
                frame,
                source_metadata,
                current_snapshot_as_of=current_snapshot_as_of if current_snapshot else None,
                current_snapshot_effective_at=(
                    current_snapshot_effective_at if current_snapshot else None
                ),
            )
        ciq_percentage_points = pd.Series(False, index=frame.index)
        if "metric" in frame.columns:
            ciq_percentage_points = (
                frame["metric"].map(_column_key).isin(CIQ_PERCENTAGE_POINT_METRICS)
            )
            if ciq_percentage_points.any():
                source_metadata["value_normalization"] = {
                    "ciq_percentage_points": "divide_by_100",
                    "metrics": sorted(CIQ_PERCENTAGE_POINT_METRICS),
                }
            frame["metric"] = frame["metric"].map(
                lambda value: CANONICAL_METRIC_ALIASES.get(_column_key(value), _column_key(value))
            )
        contract = CONTRACTS[dataset]
        issues: list[DataQualityIssue] = []
        missing = [column for column in contract.required if column not in frame.columns]
        if missing:
            issue = DataQualityIssue(
                severity=Severity.ERROR,
                dataset=dataset.value,
                code="MISSING_REQUIRED_COLUMNS",
                message=f"Missing required columns: {', '.join(missing)}",
                source_file_id=source_file_id,
            )
            self.store.record_issue(issue)
            self.store.register_source_file(
                source_file_id=source_file_id,
                dataset=dataset.value,
                original_name=source.name,
                archived_path=str(archived),
                sha256=digest,
                row_count=0,
                metadata={"status": "rejected", "missing_columns": missing},
            )
            raise ValueError(issue.message)

        for column, value in (contract.defaults or {}).items():
            if column not in frame.columns:
                frame[column] = value
        for column in contract.optional:
            if column not in frame.columns:
                frame[column] = None

        for column in contract.date_columns:
            frame[column] = pd.to_datetime(frame[column], errors="coerce").dt.date
        for column in contract.datetime_columns:
            frame[column] = pd.to_datetime(frame[column], errors="coerce", utc=True).dt.tz_localize(
                None
            )
            if column == "effective_at" and date_only_availability:
                frame[column] = (
                    frame[column].dt.normalize()
                    + pd.Timedelta(days=1)
                    - pd.Timedelta(microseconds=1)
                )

        numeric_columns = {
            "value",
            "open",
            "high",
            "low",
            "close",
            "adjusted_close",
            "volume",
            "institutional_pct",
            "institutional_change",
            "shares",
            "return_1d",
            "return_1w",
            "return_1m",
        }
        for column in numeric_columns.intersection(frame.columns):
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
        if "value" in frame.columns and ciq_percentage_points.any():
            frame.loc[ciq_percentage_points, "value"] = (
                frame.loc[ciq_percentage_points, "value"] / 100.0
            )

        required_nulls = frame[list(contract.required)].isna().any(axis=1)
        rejected_rows = int(required_nulls.sum())
        if rejected_rows:
            value_only_missing = pd.Series(False, index=frame.index)
            if dataset in {DatasetKind.FUNDAMENTALS, DatasetKind.ESTIMATES, DatasetKind.MARKET_RETURNS}:
                # CIQ returns a completely empty metric cell together with empty
                # period companions when no observation exists for a company.
                # That is ordinary cross-sectional missingness, not a malformed
                # timestamp.  Identity/provenance must still be present, and any
                # row with a numeric value but a missing period or availability
                # timestamp remains a fatal point-in-time error.
                if dataset is DatasetKind.MARKET_RETURNS:
                    observation_identity = ["company_id", "ticker", "as_of_date", "effective_at"]
                    value_only_missing = frame[["return_1d", "return_1w", "return_1m"]].isna().all(axis=1) & frame[
                        observation_identity
                    ].notna().all(axis=1)
                else:
                    observation_identity = ["company_id", "ticker", "as_of_date", "metric"]
                    value_only_missing = frame["value"].isna() & frame[
                        observation_identity
                    ].notna().all(axis=1)
            fatal_count = int((required_nulls & ~value_only_missing).sum())
            unavailable_count = int((required_nulls & value_only_missing).sum())
            if unavailable_count:
                issue = DataQualityIssue(
                    severity=Severity.WARNING,
                    dataset=dataset.value,
                    code="NULL_REQUIRED_VALUE",
                    message=(
                        f"Rejected {unavailable_count} rows whose source returned no "
                        "numeric observation"
                    ),
                    source_file_id=source_file_id,
                )
                issues.append(issue)
                self.store.record_issue(issue)
            if fatal_count:
                issue = DataQualityIssue(
                    severity=Severity.ERROR,
                    dataset=dataset.value,
                    code="NULL_REQUIRED_VALUE",
                    message=f"Rejected {fatal_count} rows with invalid required identity/time values",
                    source_file_id=source_file_id,
                )
                issues.append(issue)
                self.store.record_issue(issue)
            frame = frame.loc[~required_nulls].copy()

        cutoff = pd.to_datetime(frame["as_of_date"]).dt.normalize() + pd.Timedelta(days=1)
        future = pd.to_datetime(frame["effective_at"]) >= cutoff
        if future.any():
            count = int(future.sum())
            issue = DataQualityIssue(
                severity=Severity.ERROR,
                dataset=dataset.value,
                code="FUTURE_EFFECTIVE_TIMESTAMP",
                message=f"Rejected {count} rows whose effective_at is after as_of_date",
                source_file_id=source_file_id,
            )
            issues.append(issue)
            self.store.record_issue(issue)
            frame = frame.loc[~future].copy()
            rejected_rows += count

        now = datetime.now(timezone.utc).replace(tzinfo=None)
        frame["source_file_id"] = source_file_id
        frame["ingested_at"] = now
        selected = frame[TABLE_COLUMNS[contract.table]].copy()
        self.store.insert_frame(contract.table, selected)
        self.store.register_source_file(
            source_file_id=source_file_id,
            dataset=dataset.value,
            original_name=source.name,
            archived_path=str(archived),
            sha256=digest,
            row_count=len(selected),
            metadata={
                "status": "accepted",
                "rejected_rows": rejected_rows,
                "columns": list(selected.columns),
                "source_scope": "current_snapshot" if current_snapshot else "point_in_time_export",
                "timestamp_provenance": (
                    "operator_supplied_export_timestamp"
                    if current_snapshot
                    else "embedded_spg_formula_and_source_columns"
                    if source_metadata.get("embedded_as_of_date")
                    or source_metadata.get("historical_snapshot_dates")
                    else "source_columns"
                ),
                "source_formula_metadata": source_metadata,
                "parameter_rows_dropped": parameter_rows_dropped,
            },
        )
        return ImportResult(
            source_file_id=source_file_id,
            dataset=dataset,
            sha256=digest,
            archived_path=str(archived),
            imported_rows=len(selected),
            rejected_rows=rejected_rows,
            issues=issues,
        )


def historical_universe_gaps(store: Store, start_date, end_date) -> pd.DataFrame:
    """List historical constituent identities still missing institutional data."""
    return store.query_df(
        """
        WITH historical AS (
            SELECT company_id, ticker,
                   MIN(member_from) AS first_member,
                   MAX(COALESCE(member_to, ?)) AS last_member
            FROM index_membership
            WHERE index_code = 'SP500'
              AND member_from <= ?
              AND (member_to IS NULL OR member_to >= ?)
            GROUP BY company_id, ticker
        ), fundamental_ids AS (
            SELECT DISTINCT company_id FROM fundamentals
            WHERE as_of_date BETWEEN ? AND ?
        ), estimate_ids AS (
            SELECT DISTINCT company_id FROM estimates
            WHERE metric = 'eps_estimate' AND as_of_date BETWEEN ? AND ?
        ), instrument_ids AS (
            SELECT DISTINCT company_id FROM instruments
        )
        SELECT h.company_id, h.ticker, h.first_member, h.last_member,
               (f.company_id IS NULL) AS needs_fundamentals,
               (e.company_id IS NULL) AS needs_estimates,
               (i.company_id IS NULL) AS needs_instrument_identity
        FROM historical AS h
        LEFT JOIN fundamental_ids AS f ON f.company_id = h.company_id
        LEFT JOIN estimate_ids AS e ON e.company_id = h.company_id
        LEFT JOIN instrument_ids AS i ON i.company_id = h.company_id
        WHERE f.company_id IS NULL OR e.company_id IS NULL OR i.company_id IS NULL
        ORDER BY h.ticker, h.first_member
        """,
        [
            end_date,
            end_date,
            start_date,
            start_date,
            end_date,
            start_date,
            end_date,
        ],
    )


def certify_point_in_time(
    store: Store, start_date, end_date, required: Iterable[str] | None = None
) -> tuple[bool, list[str]]:
    minimum_company_coverage = 400
    minimum_universe_coverage = 0.90
    required = tuple(required or ("index_membership", "fundamentals", "estimates", "prices"))
    notes: list[str] = []
    coverage_failures: list[str] = []
    status = {row["dataset"]: row for row in store.source_status()}
    synthetic = store.query_df(
        "SELECT COUNT(*) AS count FROM source_files WHERE original_name LIKE ?",
        ["synthetic_%"],
    )
    synthetic_fixture = bool(
        not synthetic.empty and int(synthetic.iloc[0]["count"])
    )
    if synthetic_fixture:
        # The acceptance fixture intentionally contains only 35 companies.  It
        # must still satisfy the same 90% cross-sectional and date gates, while
        # the real study retains the institutional 400-company floor.
        minimum_company_coverage = 1
        notes.append(
            "SYNTHETIC DATA ONLY: engineering validation; not a Capital IQ investment result"
        )
    missing = [dataset for dataset in required if dataset not in status]
    if missing:
        notes.append(f"Missing authoritative datasets: {', '.join(missing)}")

    def historical_universe_coverage(
        table: str,
        *,
        frequency: str,
        label: str,
        metric: str | None = None,
    ) -> list[str]:
        if table not in {"fundamentals", "estimates"}:
            raise ValueError(table)
        expected_periods = set(pd.period_range(start_date, end_date, freq=frequency))
        metric_filter = " AND metric = ?" if metric else ""
        parameters: list[object] = [start_date, end_date]
        if metric:
            parameters.append(metric)
        parameters.extend([start_date, end_date])
        if metric:
            parameters.append(metric)
        # Aggregate at the database so the Supabase status page transfers only
        # one row per snapshot instead of tens of thousands of company rows.
        # DISTINCT in ``active`` also protects against overlapping source
        # membership intervals without inflating the denominator.
        clean = store.query_df(
            f"""
            WITH snapshot_dates AS (
                SELECT DISTINCT as_of_date AS snapshot_date
                FROM {table}
                WHERE as_of_date BETWEEN ? AND ?{metric_filter}
            ), active AS (
                SELECT DISTINCT dates.snapshot_date, member.company_id
                FROM snapshot_dates AS dates
                JOIN index_membership AS member
                  ON member.index_code = 'SP500'
                 AND member.member_from <= dates.snapshot_date
                 AND (member.member_to IS NULL OR member.member_to >= dates.snapshot_date)
            ), present AS (
                SELECT DISTINCT as_of_date AS snapshot_date, company_id
                FROM {table}
                WHERE as_of_date BETWEEN ? AND ?{metric_filter}
            )
            SELECT active.snapshot_date,
                   COUNT(*) AS active_count,
                   COUNT(present.company_id) AS covered
            FROM active
            LEFT JOIN present
              ON present.snapshot_date = active.snapshot_date
             AND present.company_id = active.company_id
            GROUP BY active.snapshot_date
            ORDER BY active.snapshot_date
            """,
            parameters,
        )
        if not clean.empty:
            clean["snapshot_date"] = pd.to_datetime(clean["snapshot_date"], errors="coerce")
            clean = clean.dropna(subset=["snapshot_date"])

        best: dict[pd.Period, tuple[int, int, float]] = {}
        for row in clean.itertuples(index=False):
            snapshot_date = pd.Timestamp(row.snapshot_date)
            period = snapshot_date.to_period(frequency)
            if period not in expected_periods:
                continue
            covered = int(row.covered)
            active_count = int(row.active_count)
            ratio = covered / active_count if active_count else 0.0
            if period not in best or ratio > best[period][2]:
                best[period] = (covered, active_count, ratio)

        weak: list[tuple[pd.Period, int, int, float]] = []
        for period in sorted(expected_periods):
            covered, active_count, ratio = best.get(period, (0, 0, 0.0))
            required_companies = max(
                minimum_company_coverage,
                math.ceil(active_count * minimum_universe_coverage),
            )
            if active_count == 0 or covered < required_companies:
                weak.append((period, covered, active_count, ratio))
        if not weak:
            return []
        preview = ", ".join(
            f"{period} ({covered}/{active}; {ratio:.1%})"
            for period, covered, active, ratio in weak[:6]
        )
        suffix = " ..." if len(weak) > 6 else ""
        return [
            f"{label} cover fewer than {minimum_universe_coverage:.0%} of the historical "
            f"S&P 500 universe in {len(weak)} {frequency.lower()} snapshot(s): "
            f"{preview}{suffix}"
        ]

    if "fundamentals" in status:
        fundamental_dates = store.query_df(
            "SELECT MIN(effective_at) AS first, MAX(effective_at) AS last FROM fundamentals"
        )
        if fundamental_dates.empty or pd.isna(fundamental_dates.iloc[0]["first"]):
            coverage_failures.append("No point-in-time fundamental observations")
        else:
            first_fundamental = pd.Timestamp(fundamental_dates.iloc[0]["first"]).date()
            last_fundamental = pd.Timestamp(fundamental_dates.iloc[0]["last"]).date()
            if first_fundamental > start_date:
                coverage_failures.append(
                    "Point-in-time fundamentals begin after the requested backtest start"
                )
            # Annual statements can legitimately be several months old at a signal date.
            # A 15-month freshness allowance catches a current-only snapshot without
            # requiring quarterly coverage from every issuer.
            minimum_recent_filing = (pd.Timestamp(end_date) - pd.DateOffset(months=15)).date()
            if last_fundamental < minimum_recent_filing:
                coverage_failures.append(
                    "Point-in-time fundamentals end too early for the requested backtest end"
                )

        coverage_failures.extend(
            historical_universe_coverage(
                "fundamentals",
                frequency="Q",
                label="Point-in-time fundamentals",
            )
        )

    if "estimates" in status:
        coverage_failures.extend(
            historical_universe_coverage(
                "estimates",
                frequency="M",
                label="Point-in-time EPS estimates",
                metric="eps_estimate",
            )
        )

    notes.extend(coverage_failures)

    error_placeholders = ",".join("?" for _ in required)
    errors = store.query_df(
        f"""
        WITH relevant_research_sources AS (
            SELECT DISTINCT 'fundamentals' AS dataset, source_file_id
            FROM fundamentals WHERE as_of_date BETWEEN ? AND ?
            UNION
            SELECT DISTINCT 'estimates' AS dataset, source_file_id
            FROM estimates WHERE as_of_date BETWEEN ? AND ?
        )
        SELECT issue.dataset, issue.code, COUNT(*) AS count
        FROM data_quality_issues AS issue
        WHERE issue.severity = 'error'
          AND issue.dataset IN ({error_placeholders})
          AND (
              issue.dataset NOT IN ('fundamentals', 'estimates')
              OR issue.source_file_id IS NULL
              OR EXISTS (
                  SELECT 1 FROM relevant_research_sources AS relevant
                  WHERE relevant.dataset = issue.dataset
                    AND relevant.source_file_id = issue.source_file_id
              )
          )
        GROUP BY issue.dataset, issue.code ORDER BY issue.dataset, issue.code
        """,
        [start_date, end_date, start_date, end_date, *required],
    )
    for row in errors.to_dict(orient="records"):
        notes.append(f"{row['dataset']}: {row['code']} ({row['count']} rows/files)")

    coverage = store.query_df(
        """
        SELECT MIN(member_from) AS first_member, MAX(member_to) AS last_member,
               SUM(CASE WHEN member_to IS NULL THEN 1 ELSE 0 END) AS active_members
        FROM index_membership WHERE index_code = 'SP500'
        """
    )
    if coverage.empty or pd.isna(coverage.iloc[0]["first_member"]):
        notes.append("No historical S&P 500 membership coverage")
    else:
        first_member = pd.Timestamp(coverage.iloc[0]["first_member"]).date()
        last_member = (
            None
            if pd.isna(coverage.iloc[0]["last_member"])
            else pd.Timestamp(coverage.iloc[0]["last_member"]).date()
        )
        if first_member > start_date:
            notes.append("Historical membership begins after the requested backtest start")
        elif (
            not int(coverage.iloc[0]["active_members"] or 0)
            and last_member
            and last_member < end_date
        ):
            notes.append("Historical membership ends before the requested backtest end")

        membership_sources = store.query_df(
            "SELECT metadata_json FROM source_files WHERE dataset = 'index_membership'"
        )
        substitutes: set[str] = set()
        for raw_metadata in membership_sources.get("metadata_json", []):
            metadata = raw_metadata
            if isinstance(metadata, str):
                try:
                    metadata = json.loads(metadata)
                except json.JSONDecodeError:
                    metadata = {}
            if isinstance(metadata, dict) and metadata.get("capital_iq_data") is False:
                substitutes.add(str(metadata.get("provider") or "unknown_public_source"))
        if substitutes:
            notes.append(
                "Membership coverage uses explicitly labelled non-Capital-IQ source(s): "
                + ", ".join(sorted(substitutes))
            )

    first_price, last_price = store.price_coverage()
    if first_price is None or last_price is None:
        notes.append("No price history")
    else:
        if first_price > start_date or last_price < end_date:
            notes.append("Price history does not span the requested backtest window")
        substitutes = [str(value) for value in store.price_sources() if str(value) != "capital_iq"]
        if substitutes:
            notes.append(
                "Price coverage uses explicitly labelled non-Capital-IQ source(s): "
                + ", ".join(substitutes)
            )

    certified = not missing and errors.empty and not coverage_failures
    certified = certified and not any(note.startswith("No ") for note in notes)
    certified = certified and not any("does not span" in note for note in notes)
    certified = certified and not any("begins after" in note for note in notes)
    if certified:
        notes.append("Point-in-time availability, membership, and price coverage checks passed")
    return certified, notes
