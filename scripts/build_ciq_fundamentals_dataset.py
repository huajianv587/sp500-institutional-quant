#!/usr/bin/env python3
"""Normalize a split-column Capital IQ point-in-time fundamentals export."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REVENUE_KEYFIELD = "329288"
PERIOD_END_KEYFIELD = "329317"
FILING_DATE_KEYFIELD = "329318"
DATE_PATTERN = re.compile(r'"(\d{1,2}/\d{1,2}/\d{4})"')


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_formula_date(formula: Any, keyfield: str) -> date | None:
    if not isinstance(formula, str) or "SPGLabel" not in formula:
        return None
    if not re.search(rf"SPGLabel\([^,]+,\s*{re.escape(keyfield)}(?:\D|$)", formula):
        return None
    matches = DATE_PATTERN.findall(formula)
    return datetime.strptime(matches[0], "%m/%d/%Y").date() if matches else None


def is_numeric(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and not (isinstance(value, float) and math.isnan(value))
    )


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def expected_snapshots() -> list[date]:
    snapshots: list[date] = []
    year, month = 2021, 9
    while (year, month) <= (2026, 6):
        next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
        snapshots.append(date.fromordinal(date(next_year, next_month, 1).toordinal() - 1))
        month += 3
        if month > 12:
            year += 1
            month -= 12
    snapshots.append(date(2026, 8, 31))
    return snapshots


def read_tickers(path: Path, expected_companies: list[str]) -> list[str]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_row = next(
            (
                row
                for row in range(1, min(sheet.max_row, 30) + 1)
                if sheet.cell(row, 1).value == "Company" and sheet.cell(row, 2).value == "Ticker"
            ),
            None,
        )
        if header_row is None:
            raise ValueError(f"{path}: List Manager Company/Ticker header not found")
        rows = [
            (
                str(sheet.cell(row, 1).value or "").strip(),
                str(sheet.cell(row, 2).value or "").split(" (", 1)[0].strip(),
            )
            for row in range(header_row + 1, sheet.max_row + 1)
            if str(sheet.cell(row, 1).value or "").strip()
        ]
        if [company for company, _ in rows] != expected_companies:
            raise ValueError(f"{path}: List Manager rows do not match the CIQ export")
        tickers = [ticker for _, ticker in rows]
        if any(not ticker for ticker in tickers):
            raise ValueError(f"{path}: every company must have a ticker")
        return tickers
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--tickers", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    args = parser.parse_args()

    formulas = load_workbook(args.input, read_only=False, data_only=False)
    values = load_workbook(args.input, read_only=False, data_only=True)
    try:
        formula_sheet = formulas[formulas.sheetnames[0]]
        value_sheet = values[values.sheetnames[0]]
        companies = [
            str(value_sheet.cell(row, 1).value or "").strip()
            for row in range(8, value_sheet.max_row + 1)
        ]
        company_ids = [
            str(value_sheet.cell(row, 2).value or "").strip()
            for row in range(8, value_sheet.max_row + 1)
        ]
        if not companies or any(not value for value in companies + company_ids):
            raise ValueError("every CIQ row must include company name and entity ID")
        tickers = read_tickers(args.tickers, companies)

        fields: dict[str, dict[date, list[Any]]] = {
            REVENUE_KEYFIELD: {},
            PERIOD_END_KEYFIELD: {},
            FILING_DATE_KEYFIELD: {},
        }
        for column in range(3, formula_sheet.max_column + 1):
            formula = formula_sheet.cell(4, column).value
            for keyfield, by_date in fields.items():
                snapshot_date = parse_formula_date(formula, keyfield)
                if snapshot_date is None:
                    continue
                if snapshot_date in by_date:
                    raise ValueError(f"duplicate keyfield {keyfield} at {snapshot_date}")
                by_date[snapshot_date] = [
                    value_sheet.cell(row, column).value for row in range(8, value_sheet.max_row + 1)
                ]

        snapshots = expected_snapshots()
        for keyfield, by_date in fields.items():
            if sorted(by_date) != snapshots:
                raise ValueError(f"keyfield {keyfield} does not contain the exact 21 snapshots")

        args.output.parent.mkdir(parents=True, exist_ok=True)
        columns = [
            "company_id",
            "ticker",
            "company_name",
            "period_end",
            "period_type",
            "effective_at",
            "as_of_date",
            "metric",
            "value",
            "unit",
        ]
        exported_rows = 0
        skipped_rows = 0
        with args.output.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            for snapshot_date in snapshots:
                for index, company_id in enumerate(company_ids):
                    revenue = fields[REVENUE_KEYFIELD][snapshot_date][index]
                    period_end = as_date(fields[PERIOD_END_KEYFIELD][snapshot_date][index])
                    filing_date = as_date(fields[FILING_DATE_KEYFIELD][snapshot_date][index])
                    if not is_numeric(revenue) or period_end is None or filing_date is None:
                        skipped_rows += 1
                        continue
                    if filing_date > snapshot_date or period_end > snapshot_date:
                        raise ValueError(
                            f"future fundamental date for {company_id} at {snapshot_date}"
                        )
                    effective_at = datetime.combine(
                        filing_date, time(23, 59, 59), tzinfo=timezone.utc
                    ).isoformat()
                    writer.writerow(
                        {
                            "company_id": company_id,
                            "ticker": tickers[index],
                            "company_name": companies[index],
                            "period_end": period_end.isoformat(),
                            "period_type": "FY",
                            "effective_at": effective_at,
                            "as_of_date": snapshot_date.isoformat(),
                            "metric": "revenue",
                            "value": revenue,
                            "unit": "USD thousands",
                        }
                    )
                    exported_rows += 1
    finally:
        formulas.close()
        values.close()

    manifest = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "company_count": len(companies),
        "snapshot_count": len(snapshots),
        "first_snapshot": snapshots[0].isoformat(),
        "last_snapshot": snapshots[-1].isoformat(),
        "exported_rows": exported_rows,
        "skipped_rows": skipped_rows,
        "keyfields": {
            "revenue": REVENUE_KEYFIELD,
            "period_end": PERIOD_END_KEYFIELD,
            "filing_date": FILING_DATE_KEYFIELD,
        },
        "inputs": [
            {"path": str(path), "sha256": sha256(path), "bytes": path.stat().st_size}
            for path in (args.input, args.tickers)
        ],
        "output": str(args.output),
        "output_sha256": sha256(args.output),
    }
    args.manifest.parent.mkdir(parents=True, exist_ok=True)
    args.manifest.write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps(manifest, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
