#!/usr/bin/env python3
"""Build a normalized point-in-time estimates dataset from CIQ Pro exports.

Capital IQ Screener exports may need to be split across workbooks because a
large set of repeated As Of Date columns is cumbersome to configure.  This
utility joins one or more EPS value workbooks to a companion fiscal-period-end
workbook and a List Manager ticker export.  It fails closed on row-order,
formula, date-coverage, or overlapping-value conflicts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

VALUE_KEYFIELD = "290476"
PERIOD_KEYFIELD = "290486"
DATE_PATTERN = re.compile(r'"(\d{1,2}/\d{1,2}/\d{4})"')


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_formula_date(formula: Any, keyfield: str) -> date | None:
    if not isinstance(formula, str) or "SPGLabel" not in formula:
        return None
    if not re.search(rf"SPGLabel\([^,]+,\s*{re.escape(keyfield)}(?:\D|$)", formula):
        return None
    matches = DATE_PATTERN.findall(formula)
    if not matches:
        return None
    return datetime.strptime(matches[0], "%m/%d/%Y").date()


def is_missing(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value))


def read_ciq_snapshot_workbook(
    path: Path, *, keyfield: str
) -> tuple[list[str], list[str], dict[date, list[Any]]]:
    # These exports are small enough to load in memory.  Normal mode avoids
    # openpyxl's expensive repeated random access on a read-only worksheet.
    formulas = load_workbook(path, read_only=False, data_only=False)
    values = load_workbook(path, read_only=False, data_only=True)
    try:
        formula_sheet = formulas[formulas.sheetnames[0]]
        value_sheet = values[values.sheetnames[0]]
        if formula_sheet.max_row < 8 or formula_sheet.max_column < 3:
            raise ValueError(f"{path}: workbook is missing the CIQ table preamble")

        companies = [
            str(value_sheet.cell(row, 1).value or "").strip()
            for row in range(8, value_sheet.max_row + 1)
        ]
        company_ids = [
            str(value_sheet.cell(row, 2).value or "").strip()
            for row in range(8, value_sheet.max_row + 1)
        ]
        if not companies or any(not value for value in companies + company_ids):
            raise ValueError(f"{path}: every data row must include company name and entity ID")

        by_date: dict[date, list[Any]] = {}
        for column in range(3, formula_sheet.max_column + 1):
            snapshot_date = parse_formula_date(formula_sheet.cell(4, column).value, keyfield)
            if snapshot_date is None:
                continue
            column_values = [
                value_sheet.cell(row, column).value for row in range(8, value_sheet.max_row + 1)
            ]
            if snapshot_date not in by_date:
                by_date[snapshot_date] = column_values
                continue
            existing = by_date[snapshot_date]
            for index, value in enumerate(column_values):
                if is_missing(value):
                    continue
                if not is_missing(existing[index]) and existing[index] != value:
                    raise ValueError(
                        f"{path}: conflicting duplicate {keyfield} values for "
                        f"company row {index + 1} at {snapshot_date}"
                    )
                existing[index] = value
        if not by_date:
            raise ValueError(f"{path}: no keyfield {keyfield} dated columns found")
        return companies, company_ids, by_date
    finally:
        formulas.close()
        values.close()


def read_tickers(path: Path, expected_companies: list[str]) -> list[str]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_row = next(
            (
                row
                for row in range(1, min(sheet.max_row, 30) + 1)
                if str(sheet.cell(row, 1).value or "").strip() == "Company"
                and str(sheet.cell(row, 2).value or "").strip() == "Ticker"
            ),
            None,
        )
        if header_row is None:
            raise ValueError(f"{path}: List Manager Company/Ticker header not found")
        rows: list[tuple[str, str]] = []
        for row in range(header_row + 1, sheet.max_row + 1):
            company = str(sheet.cell(row, 1).value or "").strip()
            raw_ticker = str(sheet.cell(row, 2).value or "").strip()
            if not company:
                continue
            ticker = raw_ticker.split(" (", 1)[0].strip()
            rows.append((company, ticker))
        companies = [company for company, _ in rows]
        if companies != expected_companies:
            raise ValueError(f"{path}: List Manager rows do not exactly match the CIQ export order")
        tickers = [ticker for _, ticker in rows]
        if any(not ticker for ticker in tickers):
            raise ValueError(f"{path}: every company must have a ticker")
        return tickers
    finally:
        workbook.close()


def merge_value_chunks(
    paths: list[Path], expected_companies: list[str], expected_ids: list[str]
) -> dict[date, list[Any]]:
    merged: dict[date, list[Any]] = {}
    for path in paths:
        companies, company_ids, chunk = read_ciq_snapshot_workbook(path, keyfield=VALUE_KEYFIELD)
        if companies != expected_companies or company_ids != expected_ids:
            raise ValueError(f"{path}: company rows differ from the period workbook")
        for snapshot_date, values in chunk.items():
            existing = merged.setdefault(snapshot_date, [None] * len(values))
            for index, value in enumerate(values):
                if is_missing(value):
                    continue
                if not is_missing(existing[index]) and existing[index] != value:
                    raise ValueError(
                        f"{path}: conflicting EPS values for company row {index + 1} "
                        f"at {snapshot_date}"
                    )
                existing[index] = value
    return merged


def expected_month_ends(start: date, end: date) -> list[date]:
    result: list[date] = []
    year, month = start.year, start.month
    while (year, month) <= (end.year, end.month):
        next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
        month_end = date.fromordinal(date(next_year, next_month, 1).toordinal() - 1)
        result.append(month_end)
        year, month = next_year, next_month
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--periods", type=Path, required=True)
    parser.add_argument("--tickers", type=Path, required=True)
    parser.add_argument("--values", type=Path, nargs="+", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--start", type=date.fromisoformat, default=date(2021, 9, 1))
    parser.add_argument("--end", type=date.fromisoformat, default=date(2026, 8, 31))
    args = parser.parse_args()

    companies, company_ids, periods = read_ciq_snapshot_workbook(
        args.periods, keyfield=PERIOD_KEYFIELD
    )
    tickers = read_tickers(args.tickers, companies)
    estimates = merge_value_chunks(args.values, companies, company_ids)

    expected_dates = expected_month_ends(args.start, args.end)
    if sorted(periods) != expected_dates:
        raise ValueError("period workbook does not contain the exact requested month ends")
    if sorted(estimates) != expected_dates:
        missing = sorted(set(expected_dates) - set(estimates))
        extra = sorted(set(estimates) - set(expected_dates))
        raise ValueError(f"EPS value chunks have missing={missing} extra={extra}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "company_id",
        "ticker",
        "company_name",
        "fiscal_period",
        "effective_at",
        "as_of_date",
        "metric",
        "value",
        "unit",
    ]
    exported_rows = 0
    skipped_missing_value = 0
    skipped_missing_period = 0
    with args.output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for snapshot_date in expected_dates:
            effective_at = datetime.combine(
                snapshot_date, time(23, 59, 59), tzinfo=timezone.utc
            ).isoformat()
            for index, company_id in enumerate(company_ids):
                value = estimates[snapshot_date][index]
                fiscal_period = periods[snapshot_date][index]
                if is_missing(value):
                    skipped_missing_value += 1
                    continue
                if is_missing(fiscal_period):
                    skipped_missing_period += 1
                    continue
                if isinstance(fiscal_period, datetime):
                    fiscal_period = fiscal_period.date()
                writer.writerow(
                    {
                        "company_id": company_id,
                        "ticker": tickers[index],
                        "company_name": companies[index],
                        "fiscal_period": fiscal_period.isoformat(),
                        "effective_at": effective_at,
                        "as_of_date": snapshot_date.isoformat(),
                        "metric": "eps_estimate",
                        "value": value,
                        "unit": "USD/share",
                    }
                )
                exported_rows += 1

    inputs = [args.periods, args.tickers, *args.values]
    manifest = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "output": str(args.output),
        "output_sha256": sha256(args.output),
        "inputs": [
            {"path": str(path), "sha256": sha256(path), "bytes": path.stat().st_size}
            for path in inputs
        ],
        "company_count": len(companies),
        "snapshot_count": len(expected_dates),
        "first_snapshot": expected_dates[0].isoformat(),
        "last_snapshot": expected_dates[-1].isoformat(),
        "exported_rows": exported_rows,
        "skipped_missing_value": skipped_missing_value,
        "skipped_missing_period": skipped_missing_period,
        "value_keyfield": VALUE_KEYFIELD,
        "period_keyfield": PERIOD_KEYFIELD,
    }
    args.manifest.parent.mkdir(parents=True, exist_ok=True)
    args.manifest.write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps(manifest, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
