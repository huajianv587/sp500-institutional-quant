from __future__ import annotations

from datetime import date, datetime

import numpy as np
import pandas as pd
import pytest

from institutional_quant.backtest import (
    BacktestEngine,
    apply_transaction_cost,
    derive_cost_sensitivity,
)
from institutional_quant.calculations import derive_fundamental_features
from institutional_quant.config import Settings
from institutional_quant.factors import FactorEngine, _combine_factor_families
from institutional_quant.ingestion import CapitalIQImporter, certify_point_in_time
from institutional_quant.ml import WalkForwardModel
from institutional_quant.portfolio import PortfolioOptimizer
from institutional_quant.schemas import BacktestResult, BacktestSpec, DatasetKind, StrategyMetrics
from institutional_quant.storage import DuckDBStore


def test_deterministic_financial_ratios() -> None:
    frame = pd.DataFrame(
        [
            {
                "net_income": 10,
                "market_cap": 100,
                "free_cash_flow": 8,
                "ebitda": 20,
                "enterprise_value": 200,
                "nopat": 9,
                "invested_capital": 90,
                "gross_profit": 40,
                "total_assets": 160,
                "operating_cash_flow": 13,
                "net_debt": 30,
                "revenue": 120,
                "revenue_prior_year": 100,
                "eps": 2.4,
                "eps_prior_year": 2,
                "operating_margin": 0.20,
                "operating_margin_prior_year": 0.18,
            }
        ]
    )
    result = derive_fundamental_features(frame).iloc[0]
    assert result.earnings_yield == pytest.approx(0.10)
    assert result.fcf_yield == pytest.approx(0.08)
    assert result.revenue_growth == pytest.approx(0.20)
    assert result.margin_change == pytest.approx(0.02)


def test_exported_valuation_multiples_are_deterministic_fallbacks() -> None:
    frame = pd.DataFrame([{"price_to_earnings": 20.0, "tev_ebitda": 10.0}])

    result = derive_fundamental_features(frame).iloc[0]

    assert result.earnings_yield == pytest.approx(0.05)
    assert result.ebitda_to_ev == pytest.approx(0.10)


def test_ciq_thousands_to_millions_scale_is_calibrated_from_reported_multiples() -> None:
    frame = pd.DataFrame(
        {
            "net_income": [10_000.0, 20_000.0, 30_000.0, -5_000.0],
            "market_cap": [100.0, 200.0, 300.0, 50.0],
            "price_to_earnings": [10.0, 10.0, 10.0, np.nan],
        }
    )

    result = derive_fundamental_features(frame)

    assert result.loc[:2, "earnings_yield"].tolist() == pytest.approx([0.1, 0.1, 0.1])
    assert result.loc[3, "earnings_yield"] == pytest.approx(-0.1)


def test_symbol_and_sector_are_point_in_time(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "test.duckdb")
    store.initialize()
    columns = [
        "company_id",
        "ticker",
        "company_name",
        "sector",
        "currency",
        "effective_at",
        "as_of_date",
        "source_file_id",
        "ingested_at",
    ]
    rows = [
        [
            "C1",
            "OLD",
            "Company",
            "Energy",
            "USD",
            datetime(2018, 1, 1),
            date(2026, 1, 1),
            "s1",
            datetime.now(),
        ],
        [
            "C1",
            "NEW",
            "Company",
            "Technology",
            "USD",
            datetime(2022, 1, 1),
            date(2026, 1, 1),
            "s1",
            datetime.now(),
        ],
    ]
    store.insert_frame("instruments", pd.DataFrame(rows, columns=columns))
    store.insert_frame(
        "index_membership",
        pd.DataFrame(
            [
                [
                    "C1",
                    "OLD",
                    "SP500",
                    date(2018, 1, 1),
                    date(2021, 12, 31),
                    datetime(2018, 1, 1),
                    date(2026, 1, 1),
                    "s1",
                    datetime.now(),
                ],
                [
                    "C1",
                    "NEW",
                    "SP500",
                    date(2022, 1, 1),
                    None,
                    datetime(2022, 1, 1),
                    date(2026, 1, 1),
                    "s1",
                    datetime.now(),
                ],
            ],
            columns=[
                "company_id",
                "ticker",
                "index_code",
                "member_from",
                "member_to",
                "effective_at",
                "as_of_date",
                "source_file_id",
                "ingested_at",
            ],
        ),
    )
    old = FactorEngine(store)._universe(date(2020, 6, 1)).iloc[0]
    new = FactorEngine(store)._universe(date(2024, 6, 1)).iloc[0]
    assert (old.ticker, old.sector) == ("OLD", "Energy")
    assert (new.ticker, new.sector) == ("NEW", "Technology")


def test_duckdb_migrates_fundamental_primary_key_without_losing_rows(tmp_path) -> None:
    import duckdb

    path = tmp_path / "legacy.duckdb"
    connection = duckdb.connect(str(path))
    connection.execute(
        """
        CREATE TABLE fundamentals (
            company_id VARCHAR NOT NULL, ticker VARCHAR NOT NULL, period_end DATE NOT NULL,
            period_type VARCHAR NOT NULL, effective_at TIMESTAMP NOT NULL,
            as_of_date DATE NOT NULL, metric VARCHAR NOT NULL, value DOUBLE, unit VARCHAR,
            source_file_id VARCHAR NOT NULL, ingested_at TIMESTAMP NOT NULL,
            PRIMARY KEY (company_id, period_end, period_type, effective_at, metric)
        )
        """
    )
    connection.execute(
        """
        INSERT INTO fundamentals VALUES (
            'C1', 'ONE', DATE '2020-12-31', 'FY', TIMESTAMP '2021-02-15 23:59:59',
            DATE '2021-08-31', 'revenue', 100, NULL, 'src', TIMESTAMP '2021-09-01'
        )
        """
    )
    connection.close()

    store = DuckDBStore(path)
    store.initialize()

    info = store.query_df("PRAGMA table_info('fundamentals')")
    primary_key = set(info.loc[info["pk"] > 0, "name"].astype(str))
    assert "as_of_date" in primary_key
    assert store.query_df("SELECT COUNT(*) AS count FROM fundamentals").iloc[0]["count"] == 1


def test_fundamental_snapshot_versions_are_visible_only_after_their_as_of_date(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "fundamental-versions.duckdb")
    store.initialize()
    base = {
        "company_id": "C1",
        "ticker": "ONE",
        "period_end": date(2020, 12, 31),
        "period_type": "FY",
        "effective_at": datetime(2021, 2, 15, 23, 59),
        "metric": "revenue",
        "unit": None,
        "source_file_id": "src",
        "ingested_at": datetime(2021, 10, 1),
    }
    snapshots = pd.DataFrame(
        [
            {**base, "as_of_date": date(2021, 8, 31), "value": 100.0},
            {**base, "as_of_date": date(2021, 9, 30), "value": 110.0},
        ]
    )[
        [
            "company_id",
            "ticker",
            "period_end",
            "period_type",
            "effective_at",
            "as_of_date",
            "metric",
            "value",
            "unit",
            "source_file_id",
            "ingested_at",
        ]
    ]
    store.insert_frame("fundamentals", snapshots)
    engine = FactorEngine(store)

    august, _ = engine._latest_metrics("fundamentals", date(2021, 8, 31))
    september, _ = engine._latest_metrics("fundamentals", date(2021, 9, 30))

    assert august.iloc[0]["revenue"] == pytest.approx(100.0)
    assert september.iloc[0]["revenue"] == pytest.approx(110.0)


def test_current_snapshot_metrics_are_not_visible_before_their_as_of_date(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "metric-asof.duckdb")
    store.initialize()
    store.insert_frame(
        "fundamentals",
        pd.DataFrame(
            [
                {
                    "company_id": "C1",
                    "ticker": "ONE",
                    "period_end": date(2025, 12, 31),
                    "period_type": "FY",
                    "effective_at": datetime(2026, 2, 1, 23, 59),
                    "as_of_date": date(2026, 9, 1),
                    "metric": "revenue",
                    "value": 100.0,
                    "unit": None,
                    "source_file_id": "current-snapshot",
                    "ingested_at": datetime(2026, 9, 1, 9, 10),
                }
            ]
        ),
    )
    engine = FactorEngine(store)

    before, _ = engine._latest_metrics("fundamentals", date(2026, 8, 31))
    on_date, _ = engine._latest_metrics("fundamentals", date(2026, 9, 1))

    assert before.empty
    assert on_date.iloc[0]["revenue"] == pytest.approx(100.0)


def test_factor_score_requires_institutional_and_total_family_coverage() -> None:
    frame = pd.DataFrame(
        [
            {
                "factor_value": None,
                "factor_quality": None,
                "factor_growth": None,
                "factor_revisions": None,
                "factor_momentum": 1.0,
                "factor_low_risk": 0.5,
            },
            {
                "factor_value": 1.0,
                "factor_quality": 0.5,
                "factor_growth": None,
                "factor_revisions": None,
                "factor_momentum": 0.25,
                "factor_low_risk": -0.25,
            },
        ]
    )

    scored = _combine_factor_families(frame)

    assert pd.isna(scored.iloc[0]["factor_score"])
    assert scored.iloc[0]["factor_family_count"] == 2
    assert scored.iloc[0]["institutional_factor_family_count"] == 0
    assert scored.iloc[1]["factor_score"] == pytest.approx(0.375)


def test_walk_forward_rejects_same_month_training() -> None:
    history = pd.DataFrame(
        {
            "as_of_date": [date(2024, 1, 31)],
            "x": [1.0],
            "factor_score": [0.0],
            "next_month_excess_return": [0.01],
        }
    )
    current = pd.DataFrame({"as_of_date": [date(2024, 1, 31)], "x": [2.0], "factor_score": [0.0]})
    with pytest.raises(ValueError, match="strictly precede"):
        WalkForwardModel().predict(history, current, ["x"])


def test_backtest_returns_remain_continuous_across_ticker_change(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "rename.duckdb")
    store.initialize()
    engine = BacktestEngine(store)
    prices = pd.DataFrame(
        [
            {
                "company_id": "C1",
                "ticker": "OLD",
                "price_date": date(2025, 1, 2),
                "open": 100.0,
                "adjusted_close": 100.0,
            },
            {
                "company_id": "C1",
                "ticker": "NEW",
                "price_date": date(2025, 2, 3),
                "open": 110.0,
                "adjusted_close": 110.0,
            },
            {
                "company_id": "BENCHMARK:SPY",
                "ticker": "SPY",
                "price_date": date(2025, 1, 2),
                "open": 100.0,
                "adjusted_close": 100.0,
            },
            {
                "company_id": "BENCHMARK:SPY",
                "ticker": "SPY",
                "price_date": date(2025, 2, 3),
                "open": 105.0,
                "adjusted_close": 105.0,
            },
        ]
    )
    forward = engine._forward_returns(
        prices,
        [date(2025, 1, 1), date(2025, 1, 31), date(2025, 2, 28)],
        "SPY",
    )[date(2025, 1, 1)]
    assert forward.loc[forward["company_id"] == "C1", "next_return"].item() == pytest.approx(0.10)

    history = engine._returns_history(
        prices,
        date(2025, 2, 3),
        pd.DataFrame({"company_id": ["C1"], "ticker": ["NEW"]}),
    )
    assert list(history.columns) == ["NEW"]
    assert history["NEW"].dropna().iloc[-1] == pytest.approx(0.10)


def test_optimizer_constraints_and_transaction_cost() -> None:
    rng = np.random.default_rng(3)
    candidates = pd.DataFrame(
        {
            "company_id": [f"C{i}" for i in range(30)],
            "ticker": [f"T{i}" for i in range(30)],
            "sector": ["A"] * 15 + ["B"] * 15,
            "ensemble_score": np.linspace(1, 0, 30),
        }
    )
    returns = pd.DataFrame(rng.normal(0, 0.01, (260, 30)), columns=candidates.ticker)
    result = PortfolioOptimizer().optimize(
        candidates, returns, date(2026, 8, 31), benchmark_sector_weights={"A": 0.5, "B": 0.5}
    )
    assert sum(position.weight for position in result.positions) == pytest.approx(1.0)
    assert max(position.weight for position in result.positions) <= 0.05 + 1e-8
    sector_a = sum(position.weight for position in result.positions if position.sector == "A")
    assert 0.42 - 1e-6 <= sector_a <= 0.58 + 1e-6
    assert apply_transaction_cost(0.02, 0.25, 10) == pytest.approx(0.01975)


def test_live_endpoint_is_rejected() -> None:
    with pytest.raises(ValueError, match="paper"):
        Settings(alpaca_paper_base_url="https://api.alpaca.markets")


def test_future_availability_is_rejected(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "future.duckdb")
    store.initialize()
    source = tmp_path / "instruments.csv"
    pd.DataFrame(
        [
            {
                "company_id": "C1",
                "ticker": "ONE",
                "company_name": "One",
                "sector": "Energy",
                "currency": "USD",
                "effective_at": "2025-02-01T00:00:00Z",
                "as_of_date": "2025-01-31",
            }
        ]
    ).to_csv(source, index=False)
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")
    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.INSTRUMENTS)
    assert result.imported_rows == 0
    assert result.rejected_rows == 1
    assert store.query_df("SELECT * FROM instruments").empty
    assert store.list_issues()[0]["code"] == "FUTURE_EFFECTIVE_TIMESTAMP"


def test_date_only_filing_availability_is_conservative_end_of_day(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "filing-date.duckdb")
    store.initialize()
    source = tmp_path / "fundamentals.csv"
    pd.DataFrame(
        [
            {
                "company_id": "C1",
                "ticker": "ONE",
                "period_end": "2024-12-31",
                "period_type": "FY",
                "financial_filing_date": "2025-01-31",
                "as_of_date": "2025-01-31",
                "metric": "total_revenue",
                "value": 100,
            }
        ]
    ).to_csv(source, index=False)
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.FUNDAMENTALS)

    assert result.imported_rows == 1
    effective_at = store.query_df("SELECT effective_at FROM fundamentals").iloc[0]["effective_at"]
    assert effective_at == pd.Timestamp("2025-01-31 23:59:59.999999")


def test_ciq_excel_preamble_header_is_detected() -> None:
    preview = pd.DataFrame(
        [
            ["=SPGTable(...)", None, None, None, None],
            ["=SPGLabel(...)", None, None, None, None],
            [None, None, None, None, None],
            [
                "SP_ENTITY_NAME",
                "SP_ENTITY_ID",
                "SP_EXCHANGE_TICKER",
                "MI_PRIMARY_INDUSTRY",
                "IQ_SECTOR",
            ],
            [
                "Example Company (NYSE:EXM)",
                "123456",
                "NYSE:EXM",
                "Example Industry",
                "Industrials",
            ],
        ]
    )

    assert CapitalIQImporter._detect_header_row(preview) == 3


def test_ciq_spg_formula_metadata_is_parsed() -> None:
    metadata = CapitalIQImporter._parse_spg_formula_metadata(
        [
            '=SPGLabel(266637,329288,"FY0","08/31/2026","Options:Curr=USD")',
            '=SPGLabel(266637,325375,"FY+1","08/31/2026","Options:Curr=USD")',
            '=SPGLabel(266637,371494,"LTM","08/31/2026","Options:Curr=USD")',
            '=SPGLabel(266637,329318,"","<>08/31/2026","Options:Curr=USD")',
        ]
    )

    assert metadata == {
        "embedded_as_of_date": "2026-08-31",
        "embedded_period_codes": ["FY+1", "FY0", "LTM"],
        "embedded_period_codes_by_keyfield": {
            "325375": ["FY+1"],
            "329288": ["FY0"],
            "371494": ["LTM"],
        },
    }


def test_ciq_multi_as_of_estimate_columns_keep_their_own_dates(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "historical_estimates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["=SPGTable(1)", None, None, None, None, None, None])
    sheet.append(
        [
            '=SPGLabel(1,267969,"")',
            '=SPGLabel(1,267961,"")',
            '=SPGLabel(1,267963,"")',
            '=SPGLabel(1,290476,"FY+1","08/31/2021")',
            '=SPGLabel(1,290486,"FY+1","08/31/2021")',
            '=SPGLabel(1,290476,"FY+1","09/30/2021")',
            '=SPGLabel(1,290486,"FY+1","09/30/2021")',
        ]
    )
    sheet.append([None] * 7)
    sheet.append([None] * 7)
    sheet.append(
        [
            "SP_ENTITY_NAME",
            "SP_ENTITY_ID",
            "SP_EXCHANGE_TICKER",
            "SP_EPS_NORM_EST",
            "SP_EPS_NORM_DATE_EST",
            "SP_EPS_NORM_EST",
            "SP_EPS_NORM_DATE_EST",
        ]
    )
    sheet.append([None, None, None, "FY+1", "FY+1", "FY+1", "FY+1"])
    sheet.append(
        [
            "Example Company",
            123456,
            "NYSE:EXM",
            5.25,
            "2022-12-31",
            5.75,
            "2022-12-31",
        ]
    )
    workbook.save(source)

    store = DuckDBStore(tmp_path / "historical-estimates.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.ESTIMATES)

    assert result.imported_rows == 2
    imported = store.query_df(
        "SELECT ticker, fiscal_period, effective_at, as_of_date, metric, value "
        "FROM estimates ORDER BY as_of_date"
    )
    assert imported["ticker"].tolist() == ["EXM", "EXM"]
    assert imported["fiscal_period"].tolist() == [
        pd.Timestamp("2022-12-31"),
        pd.Timestamp("2022-12-31"),
    ]
    assert imported["as_of_date"].tolist() == [
        pd.Timestamp("2021-08-31"),
        pd.Timestamp("2021-09-30"),
    ]
    assert imported["effective_at"].tolist() == [
        pd.Timestamp("2021-08-31 23:59:59.999999"),
        pd.Timestamp("2021-09-30 23:59:59.999999"),
    ]
    assert imported["metric"].tolist() == ["eps_estimate", "eps_estimate"]
    assert imported["value"].tolist() == pytest.approx([5.25, 5.75])


def test_ciq_current_column_is_not_relabelled_as_a_historical_snapshot(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "current_and_historical_estimates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["=SPGTable(1)", None, None, None, None, None])
    sheet.append(
        [
            '=SPGLabel(1,267961,"")',
            '=SPGLabel(1,331277,"")',
            '=SPGLabel(1,290476,"FY+1","Current")',
            '=SPGLabel(1,290476,"FY+1","08/31/2021")',
            '=SPGLabel(1,290486,"FY+1","08/31/2021")',
            None,
        ]
    )
    sheet.append([None] * 6)
    sheet.append([None] * 6)
    sheet.append(
        [
            "SP_ENTITY_ID",
            "SP_EXCHANGE_TICKER",
            "SP_EPS_NORM_EST",
            "SP_EPS_NORM_EST",
            "SP_EPS_NORM_DATE_EST",
            None,
        ]
    )
    sheet.append([None, None, "FY+1", "FY+1", "FY+1", None])
    sheet.append([123456, "NYSE:EXM", 9.0, 5.25, "2022-12-31", None])
    workbook.save(source)

    store = DuckDBStore(tmp_path / "current-and-historical.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.ESTIMATES)

    assert result.imported_rows == 1
    imported = store.query_df("SELECT as_of_date, value FROM estimates").iloc[0]
    assert imported["as_of_date"] == pd.Timestamp("2021-08-31")
    assert imported["value"] == pytest.approx(5.25)


def test_ciq_multi_as_of_fundamentals_pair_period_and_filing_date(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "historical_fundamentals.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["=SPGTable(1)", None, None, None, None, None, None, None])
    sheet.append(
        [
            '=SPGLabel(1,267961,"")',
            '=SPGLabel(1,331277,"")',
            '=SPGLabel(1,329288,"FY0","08/31/2021")',
            '=SPGLabel(1,329288,"FY0","09/30/2021")',
            '=SPGLabel(1,329317,"FY0","08/31/2021")',
            '=SPGLabel(1,329317,"FY0","09/30/2021")',
            '=SPGLabel(1,329318,"FY0","08/31/2021")',
            '=SPGLabel(1,329318,"FY0","09/30/2021")',
        ]
    )
    sheet.append([None] * 8)
    sheet.append([None] * 8)
    sheet.append(
        [
            "SP_ENTITY_ID",
            "SP_EXCHANGE_TICKER",
            "IQ_TOTAL_REV",
            "IQ_TOTAL_REV",
            "IQ_PERIOD_END",
            "IQ_PERIOD_END",
            "IQ_FINL_FILING_DATE",
            "IQ_FINL_FILING_DATE",
        ]
    )
    sheet.append([None, None, "FY0", "FY0", "FY0", "FY0", "FY0", "FY0"])
    sheet.append(
        [
            123456,
            "NYSE:EXM",
            100.0,
            110.0,
            "2020-12-31",
            "2020-12-31",
            "2021-02-15",
            "2021-02-15",
        ]
    )
    workbook.save(source)

    store = DuckDBStore(tmp_path / "historical-fundamentals.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.FUNDAMENTALS)

    assert result.imported_rows == 2
    imported = store.query_df(
        "SELECT period_end, period_type, effective_at, as_of_date, metric, value "
        "FROM fundamentals ORDER BY as_of_date"
    )
    assert imported["period_end"].tolist() == [
        pd.Timestamp("2020-12-31"),
        pd.Timestamp("2020-12-31"),
    ]
    assert imported["period_type"].tolist() == ["FY", "FY"]
    assert imported["effective_at"].tolist() == [
        pd.Timestamp("2021-02-15 23:59:59.999999"),
        pd.Timestamp("2021-02-15 23:59:59.999999"),
    ]
    assert imported["as_of_date"].tolist() == [
        pd.Timestamp("2021-08-31"),
        pd.Timestamp("2021-09-30"),
    ]
    assert imported["metric"].tolist() == ["revenue", "revenue"]
    assert imported["value"].tolist() == pytest.approx([100.0, 110.0])


def test_mixed_ciq_period_codes_are_applied_per_metric() -> None:
    frame = pd.DataFrame(
        [
            {
                "metric": "IQ_TOTAL_REV",
                "period_end": "2025-12-31",
                "effective_at": "2026-02-01",
            },
            {
                "metric": "IQ_PE",
                "period_end": "2025-12-31",
                "effective_at": "2026-02-01",
            },
            {
                "metric": "SP_MARKETCAP",
                "period_end": "2025-12-31",
                "effective_at": "2026-02-01",
            },
        ]
    )
    metadata = {
        "embedded_period_codes_by_header": {
            "iq_total_rev": ["FY0"],
            "iq_pe": ["LTM"],
        }
    }

    derived = CapitalIQImporter._derive_fundamental_period_types(
        frame,
        metadata,
        current_snapshot_as_of=date(2026, 9, 1),
        current_snapshot_effective_at=datetime(2026, 9, 1, 4, 30),
    )

    assert derived["period_type"].tolist() == ["FY", "LTM", "CURRENT"]
    assert derived.iloc[2]["period_end"] == date(2026, 9, 1)
    assert derived.iloc[2]["effective_at"] == datetime(2026, 9, 1, 4, 30)


def test_ciq_parameter_row_is_not_treated_as_an_observation() -> None:
    frame = pd.DataFrame(
        [
            {
                "company_name": None,
                "company_id": None,
                "ticker": None,
                "iq_total_rev": "FY+1",
                "period_end": "FY0",
            },
            {
                "company_name": "Example Company",
                "company_id": "123456",
                "ticker": "EXM",
                "iq_total_rev": 100,
                "period_end": "2025-12-31",
            },
        ]
    )

    cleaned, dropped = CapitalIQImporter._drop_parameter_rows(frame)

    assert dropped == 1
    assert cleaned["company_id"].tolist() == ["123456"]


def test_ciq_estimate_export_derives_target_period_and_conservative_availability(
    tmp_path,
) -> None:
    from openpyxl import Workbook

    source = tmp_path / "estimates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            None,
            None,
            None,
            '=SPGLabel(266637,325375,"FY+1","08/31/2026","Options:Curr=USD")',
            '=SPGLabel(266637,329317,"FY0","08/31/2026","Options:Curr=USD")',
        ]
    )
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    sheet.append(
        [
            "SP_ENTITY_NAME",
            "SP_ENTITY_ID",
            "SP_EXCHANGE_TICKER",
            "SP_NORM_EPS_ACT_OR_EST",
            "IQ_PERIOD_END",
        ]
    )
    sheet.append([None, None, None, "FY+1", "FY0"])
    sheet.append(["Example Company (NYSE:EXM)", 123456, "NYSE:EXM", 5.5, "2025-12-31"])
    workbook.save(source)

    store = DuckDBStore(tmp_path / "estimates.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.ESTIMATES)

    assert result.imported_rows == 1
    imported = store.query_df("SELECT * FROM estimates").iloc[0]
    assert imported["company_id"] == "123456"
    assert imported["ticker"] == "EXM"
    assert imported["fiscal_period"] == pd.Timestamp("2026-12-31")
    assert imported["effective_at"] == pd.Timestamp("2026-08-31 23:59:59.999999")
    assert imported["metric"] == "eps_estimate"
    assert imported["value"] == pytest.approx(5.5)


def test_ciq_current_estimate_export_maps_revision_counts_and_target_period(
    tmp_path,
) -> None:
    from openpyxl import Workbook

    source = tmp_path / "current_estimates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None] * 8)
    sheet.append([None] * 8)
    sheet.append([None] * 8)
    sheet.append(
        [
            '=SPGLabel(1,267969,"")',
            '=SPGLabel(1,267961,"")',
            '=SPGLabel(1,331277,"")',
            '=SPGLabel(1,290476,"FY+1","Current")',
            '=SPGLabel(1,330485,"FY+1")',
            '=SPGLabel(1,330486,"FY+1")',
            '=SPGLabel(1,330491,"FY+1")',
            '=SPGLabel(1,290486,"FY+1","Current")',
        ]
    )
    sheet.append(
        [
            "SP_ENTITY_NAME",
            "SP_ENTITY_ID",
            "SP_EXCHANGE_TICKER",
            "SP_EPS_NORM_EST",
            "SP_EPS_NORM_EST_NUM_ANALYSTS_MONTH",
            "SP_EPS_NORM_EST_UP_MONTH",
            "SP_EPS_NORM_EST_DOWN_MONTH",
            "SP_EPS_NORM_DATE_EST",
        ]
    )
    sheet.append([None, None, None, "FY+1", "FY+1", "FY+1", "FY+1", "FY+1"])
    sheet.append(["Example Company", 123456, "NYSE:EXM", 5.5, 12, 4, 2, "2027-12-31"])
    workbook.save(source)

    store = DuckDBStore(tmp_path / "current_estimates.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")
    result = CapitalIQImporter(store, settings).import_file(
        source,
        DatasetKind.ESTIMATES,
        current_snapshot_as_of=date(2026, 9, 1),
        current_snapshot_effective_at=datetime(2026, 9, 1, 10, 1),
    )

    assert result.imported_rows == 4
    imported = store.query_df("SELECT fiscal_period, metric, value FROM estimates ORDER BY metric")
    assert set(imported["metric"]) == {
        "eps_analyst_count_1m",
        "eps_estimate",
        "eps_down_revisions_1m",
        "eps_up_revisions_1m",
    }
    assert set(imported["fiscal_period"]) == {pd.Timestamp("2027-12-31")}


def test_revision_breadth_uses_analyst_coverage_and_is_bounded() -> None:
    from institutional_quant.calculations import derive_estimate_revision_features

    frame = pd.DataFrame(
        {
            "eps_analyst_count_1m": [10, 2, None],
            "eps_up_revisions_1m": [6, 5, 1],
            "eps_down_revisions_1m": [2, 0, 3],
            "eps_up_revisions_3m": [8, 0, 1],
            "eps_down_revisions_3m": [1, 5, 3],
        }
    )

    derived = derive_estimate_revision_features(frame)

    assert derived["eps_revision_1m"].tolist() == pytest.approx([0.4, 1.0, -0.5])
    assert derived["eps_revision_3m"].tolist() == pytest.approx([0.7, -1.0, -0.5])


def test_ciq_keyfield_headers_are_normalized() -> None:
    frame = pd.DataFrame(
        [
            {
                "SP_ENTITY_NAME": "Example Company (NYSE:EXM)",
                "SP_ENTITY_ID": 123456,
                "SP_EXCHANGE_TICKER": "NYSE:EXM",
                "MI_PRIMARY_INDUSTRY": "Example Industry",
                "IQ_SECTOR": "Industrials",
            }
        ]
    )

    normalized = CapitalIQImporter._normalize_identifiers(CapitalIQImporter._rename_columns(frame))

    assert list(normalized.columns) == [
        "company_name",
        "company_id",
        "ticker",
        "MI_PRIMARY_INDUSTRY",
        "sector",
    ]
    assert normalized.iloc[0]["company_id"] == "123456"
    assert normalized.iloc[0]["ticker"] == "EXM"
    assert normalized.iloc[0]["company_name"] == "Example Company"


def test_ciq_metric_aliases_feed_the_factor_vocabulary(tmp_path) -> None:
    source = tmp_path / "fundamentals.csv"
    pd.DataFrame(
        [
            {
                "SP_ENTITY_ID": "123456",
                "SP_EXCHANGE_TICKER": "NYSE:EXM",
                "IQ_PERIOD_END": "2025-12-31",
                "period_type": "FY",
                "IQ_FINL_FILING_DATE": "2026-02-01",
                "as_of_date": "2026-02-01",
                "IQ_TOTAL_REV": 100,
                "IQ_EBITDA": 20,
                "IQ_ROC": 8.5,
            }
        ]
    ).to_csv(source, index=False)
    store = DuckDBStore(tmp_path / "metric-aliases.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.FUNDAMENTALS)

    assert result.imported_rows == 3
    metrics = store.query_df("SELECT metric, value FROM fundamentals ORDER BY metric").set_index(
        "metric"
    )["value"]
    assert metrics["revenue"] == pytest.approx(100)
    assert metrics["ebitda"] == pytest.approx(20)
    assert metrics["roic"] == pytest.approx(0.085)


def test_missing_estimate_value_is_rejected_as_warning(tmp_path) -> None:
    source = tmp_path / "missing-estimate.csv"
    pd.DataFrame(
        [
            {
                "company_id": "C1",
                "ticker": "ONE",
                "fiscal_period": "2027-12-31",
                "effective_at": "2026-08-31T20:00:00Z",
                "as_of_date": "2026-08-31",
                "metric": "eps_revision_1m",
                "value": None,
            }
        ]
    ).to_csv(source, index=False)
    store = DuckDBStore(tmp_path / "missing-estimate.duckdb")
    store.initialize()
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(source, DatasetKind.ESTIMATES)

    assert result.imported_rows == 0
    assert result.rejected_rows == 1
    assert result.issues[0].severity.value == "warning"
    assert store.query_df("SELECT * FROM estimates").empty


def test_current_snapshot_requires_explicit_timestamps(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "snapshot.duckdb")
    store.initialize()
    source = tmp_path / "instruments.csv"
    pd.DataFrame(
        [
            {
                "SP_ENTITY_NAME": "Example Company (NYSE:EXM)",
                "SP_ENTITY_ID": 123456,
                "SP_EXCHANGE_TICKER": "NYSE:EXM",
                "IQ_SECTOR": "Industrials",
            }
        ]
    ).to_csv(source, index=False)
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")
    importer = CapitalIQImporter(store, settings)

    with pytest.raises(ValueError, match="Both current_snapshot"):
        importer.import_file(
            source,
            DatasetKind.INSTRUMENTS,
            current_snapshot_as_of=date(2026, 9, 1),
        )

    result = importer.import_file(
        source,
        DatasetKind.INSTRUMENTS,
        current_snapshot_as_of=date(2026, 9, 1),
        current_snapshot_effective_at=datetime(2026, 9, 1, 4, 17, 7),
    )

    assert result.imported_rows == 1
    imported = store.query_df("SELECT * FROM instruments").iloc[0]
    assert imported["company_id"] == "123456"
    assert imported["ticker"] == "EXM"
    assert imported["company_name"] == "Example Company"


def test_current_fundamentals_snapshot_is_live_only_and_keeps_filing_date(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "current-fundamentals.duckdb")
    store.initialize()
    source = tmp_path / "fundamentals.csv"
    pd.DataFrame(
        [
            {
                "SP_ENTITY_ID": 123456,
                "SP_EXCHANGE_TICKER": "NYSE:EXM",
                "IQ_PERIOD_END": "2025-12-31",
                "period_type": "FY",
                "IQ_FINL_FILING_DATE": "2026-02-01",
                "IQ_TOTAL_REV": 100,
            }
        ]
    ).to_csv(source, index=False)
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    result = CapitalIQImporter(store, settings).import_file(
        source,
        DatasetKind.FUNDAMENTALS,
        current_snapshot_as_of=date(2026, 9, 1),
        current_snapshot_effective_at=datetime(2026, 9, 1, 4, 30),
    )

    assert result.imported_rows == 1
    imported = store.query_df("SELECT * FROM fundamentals").iloc[0]
    assert imported["as_of_date"] == pd.Timestamp("2026-09-01")
    assert imported["effective_at"] == pd.Timestamp("2026-02-01 23:59:59.999999")
    source_row = store.query_df("SELECT metadata_json FROM source_files").iloc[0]
    assert '"source_scope": "current_snapshot"' in source_row["metadata_json"]


def test_current_fundamentals_snapshot_requires_source_filing_date(tmp_path) -> None:
    store = DuckDBStore(tmp_path / "missing-filing-date.duckdb")
    store.initialize()
    source = tmp_path / "fundamentals.csv"
    pd.DataFrame(
        [
            {
                "SP_ENTITY_ID": 123456,
                "SP_EXCHANGE_TICKER": "NYSE:EXM",
                "IQ_PERIOD_END": "2025-12-31",
                "period_type": "FY",
                "IQ_TOTAL_REV": 100,
            }
        ]
    ).to_csv(source, index=False)
    settings = Settings(database_backend="duckdb", raw_data_dir=tmp_path / "raw")

    with pytest.raises(ValueError, match="Financial Filing Date"):
        CapitalIQImporter(store, settings).import_file(
            source,
            DatasetKind.FUNDAMENTALS,
            current_snapshot_as_of=date(2026, 9, 1),
            current_snapshot_effective_at=datetime(2026, 9, 1, 4, 30),
        )


def test_certification_parameterizes_synthetic_source_pattern() -> None:
    class CertificationStore:
        def source_status(self):
            return []

        def price_coverage(self):
            return None, None

        def price_sources(self):
            return []

        def query_df(self, sql, parameters=None):
            if "FROM source_files" in sql:
                assert "LIKE ?" in sql
                assert parameters == ["synthetic_%"]
                return pd.DataFrame({"count": [0]})
            if "FROM data_quality_issues" in sql:
                return pd.DataFrame(columns=["dataset", "code", "count"])
            if "FROM index_membership" in sql:
                return pd.DataFrame(
                    {"first_member": [pd.NaT], "last_member": [pd.NaT], "active_members": [0]}
                )
            if "FROM prices" in sql:
                return pd.DataFrame({"first": [pd.NaT], "last": [pd.NaT]})
            raise AssertionError(f"Unexpected query: {sql}")

    certified, notes = certify_point_in_time(
        CertificationStore(),
        date(2021, 9, 1),
        date(2026, 8, 31),
    )

    assert not certified
    assert "No historical S&P 500 membership coverage" in notes


def test_certification_rejects_current_only_fundamentals_and_estimates() -> None:
    class CurrentOnlyStore:
        def source_status(self):
            return [
                {"dataset": dataset}
                for dataset in ("index_membership", "fundamentals", "estimates", "prices")
            ]

        def price_coverage(self):
            return date(2021, 9, 1), date(2026, 8, 31)

        def price_sources(self):
            return ["capital_iq"]

        def query_df(self, sql, parameters=None):
            if "original_name LIKE" in sql:
                return pd.DataFrame({"count": [0]})
            if "FROM data_quality_issues" in sql:
                return pd.DataFrame(columns=["dataset", "code", "count"])
            if "MIN(effective_at)" in sql and "FROM fundamentals" in sql:
                return pd.DataFrame(
                    {
                        "first": [pd.Timestamp("2026-08-28")],
                        "last": [pd.Timestamp("2026-08-28")],
                    }
                )
            if "SELECT DISTINCT effective_at FROM estimates" in sql:
                assert parameters == [
                    datetime(2021, 9, 1),
                    datetime(2026, 9, 1),
                ]
                return pd.DataFrame({"effective_at": [datetime(2026, 8, 31, 23, 59)]})
            if "MIN(member_from)" in sql:
                return pd.DataFrame(
                    {
                        "first_member": [date(2021, 9, 1)],
                        "last_member": [pd.NaT],
                        "active_members": [500],
                    }
                )
            if "metadata_json" in sql and "index_membership" in sql:
                return pd.DataFrame(columns=["metadata_json"])
            if "MIN(price_date)" in sql:
                return pd.DataFrame({"first": [date(2021, 9, 1)], "last": [date(2026, 8, 31)]})
            if "SELECT DISTINCT source FROM prices" in sql:
                return pd.DataFrame({"source": ["capital_iq"]})
            raise AssertionError(f"Unexpected query: {sql}")

    certified, notes = certify_point_in_time(
        CurrentOnlyStore(),
        date(2021, 9, 1),
        date(2026, 8, 31),
    )

    assert not certified
    assert "Point-in-time fundamentals begin after the requested backtest start" in notes
    assert any("estimates are missing 59 monthly signal snapshot" in note for note in notes)


def test_cost_sensitivity_reuses_frozen_turnover() -> None:
    empty_metric = {
        "cagr": 0,
        "annualized_volatility": 0,
        "sharpe_zero_rf": 0,
        "sortino_zero_rf": 0,
        "max_drawdown": 0,
        "beta": 0,
        "information_ratio": 0,
        "average_one_way_turnover": 0.2,
        "monthly_hit_rate": 0.5,
        "observations": 2,
    }
    base = BacktestResult(
        spec=BacktestSpec(),
        certified_point_in_time=True,
        certification_notes=[],
        metrics=[
            StrategyMetrics(strategy="SPY buy-and-hold", **empty_metric),
            StrategyMetrics(strategy="factor_only", **empty_metric),
        ],
        monthly_returns=[
            {"date": "2021-09-01", "spy": 0.01, "factor_only": 0.02, "factor_only_turnover": 0.2},
            {"date": "2021-10-01", "spy": -0.01, "factor_only": 0.01, "factor_only_turnover": 0.4},
        ],
        factor_ic=[],
    )
    cheaper = derive_cost_sensitivity(base, 5)
    assert cheaper.spec.transaction_cost_bps == 5
    assert cheaper.monthly_returns[0]["factor_only"] == pytest.approx(0.0201)
